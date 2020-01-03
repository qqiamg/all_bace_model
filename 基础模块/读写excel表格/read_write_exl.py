import os
import openpyxl


class RWExcel(object):
    def read_excel(self, excel_name):
        '''
        读取excel操作（复制到自身代码操作）
        :param excel_name: excel的名字
        :return: 
        '''
        name = excel_name
        wb = openpyxl.load_workbook(name)
        ws = wb.active
        rows = ws.rows
        count = 0
        start = 1  # 通过调整 start 的值进行选择第几行excel开始读取, 1 为第二行开始读取
        for j, row in enumerate(rows):
            if count < start:
                count += 1
                continue
            print(j)
            print(row[0].value)  # row[0].value 为当前行第一个单元格的值

    def excel_style(self):
        '''
        excel相关的样式使用。(待补充)
        :return: 
        '''
        pass

    def write_excel_list(self, file_name, message_list, title_list=None):
        '''
        每次写一行一个列表，一个单元格一个数据,自动换行，样式需后期写入
        :param file_name: 文件的名字
        :param message_list: 需要写入的列表信息
        :param title_list: 标题的列表(可不传)
        :return: 
        '''
        if not os.path.exists('{}.xlsx'.format(file_name)):
            wb_ = openpyxl.Workbook()
            ws_ = wb_.active
            if title_list:
                title = title_list
                ws_.append(title)
            wb_.save('{}.xlsx'.format(file_name))
        else:
            pass
        mes_list = message_list
        wb = openpyxl.load_workbook('{}.xlsx'.format(file_name))
        ws = wb.active
        ws.append(mes_list)
        wb.save('{}.xlsx'.format(file_name))

    def write_excel_cell(self, file_name, row_num, message_list, title_list=None):
        '''
        逐个单元格写入excel 可以写入格式
        :param file_name: 文件名字
        :param message_list: 一行的信息列表
        :param row_num: 第几行
        :param title_list: 标题列表
        :return: 
        '''
        title_name = title_list
        zimu = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
        if not os.path.exists('{}.xlsx'.format(file_name)):
            wb = openpyxl.Workbook()
            ws = wb.active
            if title_list:  # 有标题就写标题
                if len(title_name) <= 26:  # 处理标题的长度对应的字母
                    title_zm = zimu[:len(title_name)]
                    # print(title_zm)
                else:
                    else_zm_len = len(title_name) - 26
                    title_zm = zimu
                    num = 0
                    for f_zm in zimu:
                        if num >= else_zm_len:
                            break
                        for s_zm in zimu:
                            if num >= else_zm_len:
                                break
                            ex_zm = f_zm + s_zm
                            title_zm.append(ex_zm)
                            num += 1
                for j, one_title in enumerate(title_name):  # 写标题
                    ws['{}1'.format(title_zm[j])] = one_title
            wb.save('{}.xlsx'.format(file_name))
        else:
            pass
        mes_list = message_list
        wb = openpyxl.load_workbook('{}.xlsx'.format(file_name))
        ws = wb.active
        if len(mes_list) <= 26:  # 处理标题的长度对应的字母
            mes_zm = zimu[:len(mes_list)]
            # print(mes_zm)
        else:
            else_zm_len = len(mes_list) - 26
            mes_zm = copy(zimu)
            num = 0
            for f_zm in zimu:
                if num >= else_zm_len:
                    break
                for s_zm in zimu:
                    if num >= else_zm_len:
                        break
                    ex_zm = f_zm + s_zm
                    mes_zm.append(ex_zm)
                    num += 1
        for j, one_mes in enumerate(mes_list):  # 写单元格信息
            ws['{}{}'.format(mes_zm[j], row_num)] = one_mes
        wb.save('{}.xlsx'.format(file_name))


if __name__ == '__main__':
    a = RWExcel()
    title_name = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                  'U',
                  'V', 'W', 'X', 'Y', 'Z', 'qq', 'ddd']
    mes_name = ['A1', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                'U',
                'V', 'W', 'X', 'Y', 'Z', 'qq', 'ddd']
    for i in range(2, 5):
        a.write_excel_cell('测试', i, mes_name, title_name)
        # file_name = '测试'
        # title = ['aa', 'aaaa', 'ssss']
        # list = ['a', 'q', 'e']
        # for i in range(5):
        #     a.write_excel_list(file_name, list, title)
