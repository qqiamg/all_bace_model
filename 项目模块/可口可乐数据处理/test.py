import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
import openpyxl
from itertools import product
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from copy import copy

# t = pd.read_excel('.\\test.xlsx', sheet_name='Sheet1', header=None)
# print(t)
# t2 =t[12:15]
# tx = t2[1]/100
# t2["new"] = tx
# t2[1] = t2["new"]
# del t2['new']
# print(t2)
def get_hs_mini(table):
    """处理Availability.HS.src 和 Availability.Mini.src"""
    net_list = []
    t2 = list(table[0])
    # print(t2)
    for j, i in enumerate(t2):  # 算是哪里切割
        # print(i)
        try:
            if 'NET' in i:
                print(j, i)
                net_list.append(j)
        except:
            pass
    all_title_name = table.columns.values  # 获取所有表头
    print(len(all_title_name))
    for i in range(1, len(all_title_name)):
        all_dd = []
        print('NET - Availability-Sparkling / 汽水铺货率')
        # print(net_list)
        first_net = table[net_list[0] + 1:net_list[1]]
        # print(first_net)
        sparkling = round(first_net[i].mean(), 4)
        print(sparkling)
        print('NET - Availability-STILL / 非汽水铺货率')
        second_net = table[net_list[1] + 1:len(t2) + 1]
        # print(second_net)
        STILL = round(second_net[i].mean(), 4)
        print(STILL)
        all_dd.append(first_net)
        all_dd.append(second_net)
        lins_df = pd.concat(all_dd, ignore_index=True)
        # print(lins_df)
        all_average = round(lins_df[1].mean(), 4)
        print(all_average)
        table.iloc[net_list[0], i] = sparkling
        table.iloc[net_list[1], i] = STILL
        table.iloc[2, i] = all_average
        # print(t)
    return table
    # t.to_excel('dddddddd.xlsx')


# z = get_hs_mini(t)
# print(z)
# if 'NET - Availabilit' in i:
#     print(j)
# for i in range(len(t2)):
#     lins_list = list(t2.iloc[i])
#     # print(lins_list)
#     for j, one_data in enumerate(lins_list):
#         if str(one_data) == 'nan':
#             lins_list[j] = ''
#     lins_list[1] = lins_list.pop(-1)
#     print(lins_list)

# print(tx[all_title_name])
# for i in range(tx):
#     t3 = tx[all_title_name].iloc[i]

# a = t2['new'].mean()


# print(t3)
# all_title_name = t.columns.values
# print(all_title_name)
# t2 = t[10:13]
# print(t2)
# # t2["id_"] = "ssss"
# print(t2)
# writer = pd.ExcelWriter("1.xlsx")
# # t2.to_csv("1.csv",index=False,mode="a")
# t2.to_excel(writer, index=False, sheet_name="111")
# t2.to_excel(writer, index=False, sheet_name="1333")
# writer.save()
# writer.close()
# df_list = []
# start_row = 13
# end_row = 15
# tid = 5
#
# import os
#
#
# class T(object):
#     def __init__(self):
#         self.path_list = []
#
#     def get_bybanner_city(self):
#         """获取bybanner city"""
#         lins_df = pd.read_excel('工作簿2.xlsx')
#         lins_df['tag'] = lins_df['BG'] + lins_df['OU'] + lins_df['City']
#         lins_df.drop_duplicates(subset='tag', keep='first', inplace=True)
#         lins_df.sort_values(by=["BG", "OU", "City"], inplace=True)
#         b = lins_df.iloc[:, :3]
#         print(b)
#         b.to_excel('测试去重excel.xlsx', sheet_name='测试', index=False)
#
#     def get_allfile_path(self, sPath):
#         # 列出当前路径下的所有文件夹和文件　并进行遍历
#         for schild in os.listdir(sPath):
#             # 拼接地址
#             sChildPath = os.path.join(sPath, schild)
#             # 判断当前遍历到的是文件还是文件夹
#             if os.path.isdir(sChildPath):
#                 # 再次递归调用
#                 self.get_allfile_path(sChildPath)
#             else:
#                 if '%' in sChildPath or '~$' in sChildPath:  # 除去不需要的
#                     pass
#                 else:
#                     sChildPath_list = sChildPath.rsplit("\\", 1)
#                     sChildPath_list.append(sChildPath)
#                     self.path_list.append(sChildPath_list)
#
#     def save(self):
#         bace_path = r'D:\杨伟强\工作\可口可乐数据处理\table'
#         self.get_allfile_path(bace_path)
#         # writer = pd.ExcelWriter("测试路径总和.xlsx")
#         df = pd.DataFrame(data=self.path_list, columns=["PathName", "FileName", "all_path"])
#         df.sort_values(by=["PathName", "FileName"], inplace=True)
#         print(df)
#         # df.to_excel(writer, index=False, sheet_name="Dirs")
#         df.to_excel('测试路径总和.xlsx', sheet_name='测试', index=False)
#         # writer.save()
#         # writer.close()
#
#     def get_bybanner_no(self):
#         """获取bybanner_NO"""
#         lins_df = pd.read_excel('测试路径总和.xlsx')
#         # print(lins_df)
#         writer = pd.ExcelWriter("测试banner_ou总和.xlsx")
#         banner_list = []
#         ou_list = []
#         banner_no = 1
#         ou_no = 1
#         for one in lins_df.iterrows():
#             if "by banner" in one[1]['PathName']:  # 为bybanner的路径
#                 type = one[1]['FileName'].replace('可乐-Banner-', '').replace('(Abs).xlsx', '')
#                 lins_list = []
#                 lins_list.append(one[1]["all_path"])  # 文件路径
#                 if 'base=BG' in one[1]['PathName']:  # BG 分类的
#                     lins_list.append(type)
#                     lins_list.append('Total')
#                     lins_list.append('Total')
#                     lins_list.append(banner_no)
#                 elif 'base=City' in one[1]['PathName']:  # City 分类的
#                     lins_list.append('Total')
#                     lins_list.append('Total')
#                     lins_list.append(type)
#                     lins_list.append(banner_no)
#                 else:  # OU 分类的
#                     lins_list.append('Total')
#                     lins_list.append(type)
#                     lins_list.append('Total')
#                     lins_list.append(banner_no)
#                 banner_list.append(lins_list)
#                 banner_no += 1
#             elif "by bottler" in one[1]['PathName']:
#
#                 lins_list = []
#                 lins_list.append(one[1]["all_path"])  # 文件路径
#                 if 'base=Channel' in one[1]['PathName']:  # BG 分类的
#                     type = one[1]['FileName'].replace('可乐-BG&OU-', '').replace('(Abs).xlsx', '')
#                     lins_list.append(type)
#                     lins_list.append('Total')
#                     lins_list.append(ou_no)
#                 elif 'base=Banner' in one[1]['PathName']:  # City 分类的
#                     type = one[1]['FileName'].replace('可乐-BG&OU-', '').replace('(Abs).xlsx', '').split('_')
#                     lins_list.append(type[0])
#                     lins_list.append(type[1])
#                     lins_list.append(ou_no)
#                 ou_list.append(lins_list)
#                 ou_no += 1
#
#         df_banner = pd.DataFrame(data=banner_list, columns=["Source File", "BG", "OU", "City", "NO."])
#         df_banner.to_excel(writer, index=False, sheet_name="bybanner")
#         df_ou = pd.DataFrame(data=ou_list, columns=["Source File", "Channel", "Banner", "NO."])
#         df_ou.to_excel(writer, index=False, sheet_name="byou")
#         writer.save()
#         writer.close()
#         # df.sort_values(by=["Source File"], inplace=True)
#         print(df_banner)
#
#
# if __name__ == '__main__':
#     # get_bynner_no()
#     # bace_path = r'D:\杨伟强\工作\可口可乐数据处理\table'
#     # a = T()
#     # a.get_bybanner_no()
#     # a.save()
#     # a.get_bybanner_city()
#     # a.get_allfile_path(bace_path)
#     # print(a.path_list)
#     # a = r'D:\杨伟强\工作\可口可乐数据处理\table\by banner\base=BG\可乐-Banner-CBL(%).xlsx'
#     # z = a.rsplit("\\",1)
#     # print(z)
#     lins_df = pd.read_excel('测试去重excel.xlsx')
#     # print(lins_df)
#     bace_list = ['Total', 'Total', 'Total']
#     all_list = []
#     cityNo = len(list(lins_df["City"]))
#
#     for one_df in lins_df.iterrows():
#         now_list = [one_df[1]["BG"], one_df[1]["OU"], one_df[1]["City"]]
#         # print(now_list)
#         all_list.append(now_list)
#     j = 2
#     ouNo = 1
#     a_ouNo = 1
#     maxFltNo = 1
#     city_flt = 1
#     all_fit_list = []
#     print(len(all_list))
#     for i in range(len(all_list)):
#         lins_list = ['Total', 'Total', all_list[i][2]]
#         com_name = 'Total' + 'Total' + all_list[i][2]
#         lins_list.append(com_name)
#         lins_list.append(maxFltNo)
#         city_flt = maxFltNo
#         all_fit_list.append(lins_list)
#         j += 1
#
#         if city_flt > 1:
#             lins_list = ['Total', all_list[i][1]]
#             print(i)
#             if all_list[i - 1][1] != all_list[i][1]:
#                 lins_list.append('Total')
#                 com_name = 'Total' + all_list[i][1] + 'Total'
#                 lins_list.append(com_name)
#                 try:
#                     if all_list[i + 1][1] == all_list[i][1]:
#                         maxFltNo += 1
#                 except:
#                     pass
#                 lins_list.append(maxFltNo)
#                 all_fit_list.append(lins_list)
#                 a_ouNo = maxFltNo
#                 ouNo += 1
#                 j += 1
#                 lins_list = ['Total', all_list[i][1], all_list[i][2]]
#                 com_name = 'Total' + all_list[i][1] + all_list[i][2]
#                 lins_list.append(com_name)
#                 lins_list.append(city_flt)
#                 all_fit_list.append(lins_list)
#             else:
#                 lins_list.append(all_list[i][2])
#                 com_name = 'Total' + all_list[i][1] + all_list[i][2]
#                 lins_list.append(com_name)
#                 lins_list.append(city_flt)
#                 all_fit_list.append(lins_list)
#             j += 1
#             if all_list[i - 1][0] != all_list[i][0]:
#                 lins_list = [all_list[i][0], 'Total', 'Total']
#                 com_name = all_list[i][0] + 'Total' + 'Total'
#                 lins_list.append(com_name)
#                 try:
#                     if all_list[i + 1][0] == all_list[i][0]:
#                         maxFltNo += 1
#                 except:
#                     pass
#                 lins_list.append(maxFltNo)
#                 all_fit_list.append(lins_list)
#                 j += 1
#                 # lins_list = [all_list[i][0], all_list[i][1], 'Total']
#             if all_list[i - 1][1] != all_list[i][1]:
#                 lins_list = [all_list[i][0], all_list[i][1], 'Total']
#                 com_name = all_list[i][0] + all_list[i][1] + 'Total'
#                 lins_list.append(com_name)
#                 try:
#                     if all_list[i+1][1] == all_list[i][1]:
#                         lins_list.append(a_ouNo-1)
#                     else:
#                         lins_list.append(city_flt)
#                 except:
#                     lins_list.append(city_flt)
#                 all_fit_list.append(lins_list)
#                 j+=1
#                 lins_list = [all_list[i][0], all_list[i][1], all_list[i][2]]
#                 com_name = all_list[i][0] + all_list[i][1] + all_list[i][2]
#                 lins_list.append(com_name)
#                 lins_list.append(city_flt)
#                 all_fit_list.append(lins_list)
#             else:
#                 ins_list = [all_list[i][0], all_list[i][1], all_list[i][2]]
#                 com_name = all_list[i][0] + all_list[i][1] + all_list[i][2]
#                 lins_list.append(com_name)
#                 lins_list.append(city_flt)
#                 all_fit_list.append(lins_list)
#         maxFltNo +=1
#     print(all_fit_list)
#     # print(first_list)
#     # for one_df in lins_df.iterrows():
#     #     # if 'Total' == one_df[1]["BG"] and 'Total' == one_df[1]["OU"] and 'Total' == one_df[1]["City"]:
#     #     #     pass
#     #     # else:
#     #     #     now_list = [one_df[1]["BG"], one_df[1]["OU"], one_df[1]["City"]]
#     #     #     print(now_list)
#     #     #     # one = [bace_list[0],now_list[0]]
#     #     #     # two = [bace_list[1], now_list[1]]
#     #     #     # three = [bace_list[2], now_list[2]]
#     #     #     # for o_one in one:
#     #     #     #     for o_two in two:
#     #     #     #         for o_three in three:
#     #     #     #             bace_list_  = [o_one,o_two,o_three]
#     #     #     #             print(bace_list_)
#     #     #     # print(11111)
#     #     #     for item in product([bace_list[0], now_list[0]], [bace_list[1], now_list[1]], [bace_list[2], now_list[2]]):
#     #     #         if item[0] == "Total" and item[1] == "Total" and item[2] == "Total":
#     #     #             continue
#     #     #         elif item[0] != "Total" and item[1] == "Total" and item[2] != "Total":
#     #     #             continue
#     #     #         print(item)
#     #     now_list = [one_df[1]["BG"], one_df[1]["OU"], one_df[1]["City"]]
#     #     print(now_list)
#     #     all_list.append(now_list)
#     # print(all_list)
#     # first_dict = dict()
#     # for one_first in first_list:    #一级分类
#     #     lins_list = []
#     #     if one_first == "Total":
#     #         for one in all_list:
#     #             if "Total" in lins_list:
#     #                 pass
#     #             else:
#     #                 lins_list.append("Total")
#     #             lins_list.append(one[1])
#     #     else:
#     #         for one in all_list:
#     #             if "Total" in lins_list:
#     #                 pass
#     #             else:
#     #                 lins_list.append("Total")
#     #             if one[0] == one_first:
#     #                 lins_list.append(one[1])
#     #     first_dict[one_first] = lins_list
#     # b = lins_df["OU"]
#     # second_list = list(set(b))
#     # print(second_list)
#
#     # for j, one_now_list in enumerate(now_list):
#     #     for z ,one_bace_list in enumerate(bace_list):

# colors = ['EEE5DE', 'FFE1FF', 'C1FFC1', 'C6E2FF', 'EEC591']
#
# wb = openpyxl.load_workbook('测试汇总设置表.xlsx')
# ws = wb['TableCopy']
# # sheets1=wb.sheetnames()#获取sheet页
# rows = ws.rows
# all_data_list = []
# for i in rows:
#     lins_list = []
#     for one_i in i:
#         lins_list.append(one_i.value)
#     print(lins_list)
#     all_data_list.append(lins_list)
#     # print(i)
# wb = openpyxl.load_workbook('测试fiter_bybanner1.xlsx')
# sheet = wb.create_sheet('TableCopy')
# ws = wb['TableCopy']
# for i in all_data_list:
#     ws.append(i)
# fill = PatternFill("solid", fgColor="8470FF")
# zm = ['A', 'B', 'C', 'D', 'E', 'F']
# for one_zm in zm:
#     ws['{}1'.format(one_zm)].fill = fill
# # color = colors[0]
# color_id = 0
# for i in range(2, len(all_data_list)):
#     print(colors[color_id])
#     fill = PatternFill("solid", fgColor=colors[color_id])
#     # print(fill)
#     ws['A{}'.format(i)].fill = fill
#     ws['B{}'.format(i)].fill = fill
#     ws['C{}'.format(i)].fill = fill
#     ws['D{}'.format(i)].fill = fill
#     ws['E{}'.format(i)].fill = fill
#     ws['F{}'.format(i)].fill = fill
#     if ws['A{}'.format(i)].value != ws['A{}'.format(i + 1)].value:
#         color_id += 1
#         if color_id == 5:
#             color_id = 1
#
# wb.save('{}.xlsx'.format('测试fiter_bybanner1'))

# def replace_xls(src_file, tag_file, sheet_name):
#     #        src_file是源xlsx文件，tag_file是目标xlsx文件，sheet_name是目标xlsx里的新sheet名称
#
#     print("Start sheet %s copy from %s to %s" % (sheet_name, src_file, tag_file))
#     wb = load_workbook(src_file)
#     wb2 = load_workbook(tag_file)
#
#     ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
#     ws2 = wb2.create_sheet(sheet_name.decode('utf-8'))
#
#     max_row = ws.max_row  # 最大行数
#     max_column = ws.max_column  # 最大列数
#
#     wm = zip(ws.merged_cells)  # 开始处理合并单元格
#     if len(wm) > 0:
#         for i in range(0, len(wm)):
#             cell2 = str(wm[i]).replace('(<MergeCell ', '').replace('>,)', '')
#             print("MergeCell : %s" % cell2)
#             ws2.merge_cells(cell2)
#
#     for m in range(1, max_row + 1):
#         ws2.row_dimensions[m].height = ws.row_dimensions[m].height
#         for n in range(1, 1 + max_column):
#             if n < 27:
#                 c = chr(n + 64).upper()  # ASCII字符,chr(65)='A'
#             else:
#                 if n < 677:
#                     c = chr(divmod(n, 26)[0] + 64) + chr(divmod(n, 26)[1] + 64)
#                 else:
#                     c = chr(divmod(n, 676)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[0] + 64) + chr(
#                         divmod(divmod(n, 676)[1], 26)[1] + 64)
#             i = '%s%d' % (c, m)  # 单元格编号
#             if m == 1:
#                 #				 print("Modify column %s width from %d to %d" % (n, ws2.column_dimensions[c].width ,ws.column_dimensions[c].width))
#                 ws2.column_dimensions[c].width = ws.column_dimensions[c].width
#             try:
#                 getattr(ws.cell(row=m, column=c), "value")
#                 cell1 = ws[i]  # 获取data单元格数据
#                 ws2[i].value = cell1.value  # 赋值到ws2单元格
#                 if cell1.has_style:  # 拷贝格式
#                     ws2[i].font = copy(cell1.font)
#                     ws2[i].border = copy(cell1.border)
#                     ws2[i].fill = copy(cell1.fill)
#                     ws2[i].number_format = copy(cell1.number_format)
#                     ws2[i].protection = copy(cell1.protection)
#                     ws2[i].alignment = copy(cell1.alignment)
#             except AttributeError as e:
#                 print("cell(%s) is %s" % (i, e))
#                 continue
#
#     wb2.save(tag_file)
#     wb2.close()
#     wb.close()
#
# if 'Super' == ('Hyper' or 'H/S' or 'Super' or 'Mini'):
#     print('1111')

import requests
# url = 'http://www.okooo.com/soccer/match/1072975/odds/'
# else_title = ['CBL', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 'SCCL', None, None, None, None, None, None, None, None, None, None, None, None, None]
# else_index = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK']
# merge_flag = False  #合并标志
# for j, one_else_title in enumerate(else_title):
#     if one_else_title:
#         if merge_flag == False:
#             lins_list = []
#             lins_list.append(else_index[j])
#             merge_flag = True
#         else:
#             print(lins_list)
#             # ws.merge_cells('{}1:{}1'.format(lins_list[0], lins_list[-1]))  # 合并其他单元格
#             merge_flag = False
#             lins_list = []
#             lins_list.append(else_index[j])
#     else:
#         lins_list.append(else_index[j])
#     if j == len(else_title)-1:
#         print(lins_list)

# zimu = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U','V', 'W', 'X', 'Y', 'Z']
# max_lie = 81
# # 生成索引
# if max_lie <= 26:  # 处理标题的长度对应的字母
#     mes_zm = zimu[:max_lie]
#     # print(mes_zm)
# else:
#     else_zm_len = max_lie - 26
#     mes_zm = copy(zimu)
#     num = 0
#     for f_zm in zimu:
#         print(f_zm)
#         if num >= else_zm_len:
#             break
#         for s_zm in zimu:
#             print(s_zm)
#             if num >= else_zm_len:
#                 break
#             ex_zm = f_zm + s_zm
#             mes_zm.append(ex_zm)
#             num += 1
# print(mes_zm)
# wb = openpyxl.load_workbook('测试设置表.xlsx')
# a = wb.sheetnames
# print(a)

# nowday = datetime.date.today()
# print(type(nowday))
# one_day = datetime.timedelta(days=1)
# last_week_list = []
# for i in range(1, 8):
#     nowday -= one_day

# save_path = './download_file/stroe_files/'
# fitter_table = 'SKU Availability by Store'
# wb = openpyxl.load_workbook(save_path + '{}.xlsx'.format(fitter_table))
# ws = wb.active
# all_row = ws.max_column  # 获取最大行数
# print('开始')
# print(all_row)
# all_data_list = []
# rows = ws.rows
# count = 0
# if self.fitter_table == 'SOVI by Store':
#     start = 3  # 通过调整 start 的值进行选择第几行excel开始读取, 1 为第二行开始读取
# else:
#     start = 2
# for j, row in enumerate(rows):
#     if count < start:
#         count += 1
#         continue
#     lins_list = []
#     for i in range(all_row):
#         # print(row[0].value)
#         lins_list.append(row[i].value)
#     print(lins_list)

