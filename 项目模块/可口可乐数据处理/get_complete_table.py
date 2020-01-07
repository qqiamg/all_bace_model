import os
import sys
import time
import datetime
import pandas as pd
import openpyxl
import sqlite3
from openpyxl.styles import PatternFill, Alignment, Font
from concurrent.futures import ThreadPoolExecutor

from PyQt5 import QtWidgets
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtWidgets import (QTableWidgetItem)
from PyQt5.QtCore import QThread, pyqtSignal, Qt

from main_windows import Ui_Form
from wait_windows import Wait_Form  # 等待界面

##生成完整的结果表并展示##

"""
待处理：
处理0的问题(完成)
处理部分百分号问题（完成）
弄转置表(完成)
数据生成：存进数据库的， 和转置表的
"""

all_banner_data = dict()
all_ou_data = dict()
all_table = ['Total score.Hyper', 'Total score.Super', 'Total score.Mini',
             'Availability.HS', 'Availability.Mini', 'SOVI', 'Cooler', '2nd Display',
             'Thematic Display', 'Price']
transpose_table_list = ['SKU Availability by Store', 'SKU Facing by Store', 'SKU Price by Store',
                        'SOVI by Store', 'Cooler by Store', '2nd Display by Store', 'Thematic by Store']


class GenerateTable(QThread):
    """生成完整的table(下载)"""
    str_out = pyqtSignal(str, str)  # 打印窗口信号

    def __init__(self, fitter_type, fitter_table, fitter_list):
        super().__init__()
        self.setting_file = '测试设置表.xlsx'
        self.bannerS_file = 'bybannerSS.xlsx'
        self.ouS_file = 'byouSS.xlsx'
        self.file_title_path = './stc_title/'  # 表头路径
        self.fitter_type = fitter_type  # 表类型
        # banner测试
        self.fitter_list = fitter_list  # 筛选项
        self.fitter_table = fitter_table  # 表分类
        self.table_index_list = []  # 获取到的表的id
        self.download_path = './download_file/'
        self.store_path = './download_file/stroe_files/'
        if not os.path.exists(self.download_path):  # 新建文件夹
            os.makedirs(self.download_path)

    def get_banner_fitter_id(self):
        """获取banner 的选项对应的id"""
        start = time.time()
        banner_pd = pd.read_excel(self.setting_file, sheet_name='fiterbybanner')
        # print(bananer_pd)
        banner_all_list = []  # 所有banner 的选项id
        for one_row in banner_pd.iterrows():
            # print(one_row[1])
            banner_all_list.append([one_row[1]['BG'], one_row[1]['OU'], one_row[1]['City'], one_row[1]['FilterNo']])
        print('banner筛选长度(显示部分数据)：' + str(len(banner_all_list)))
        print(banner_all_list[:6])
        end = time.time()
        print('获取banner 的选项对应的id 时间：%.2f秒' % (end - start))
        return banner_all_list

    def get_ou_fitter_id(self):
        banner_pd = pd.read_excel(self.setting_file, sheet_name='fiterbyou')
        # print(bananer_pd)
        ou_all_list = []  # 所有ou 的选项id
        for one_row in banner_pd.iterrows():
            # print(one_row[1])
            ou_all_list.append([one_row[1]['Channel'], one_row[1]['Banner'], one_row[1]['FilterNo']])
        print('ou筛选长度(显示部分数据)：' + str(len(ou_all_list)))
        print(ou_all_list[:6])
        return ou_all_list

    def get_fitter_table(self):
        """根据对应的id找到对应的表"""
        zimu = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
        fitter_id = ''
        start = time.time()
        if self.fitter_type == 'bybanner':
            table_fitter_pd = all_banner_data[self.fitter_table]
            fitter_all_list = self.get_banner_fitter_id()  # 获取所有筛选
        else:  # ou表
            # table_fitter_pd = pd.read_excel(self.ouS_file, sheet_name=self.fitter_table + '.src')
            table_fitter_pd = all_ou_data[self.fitter_table]
            fitter_all_list = self.get_ou_fitter_id()  # 获取所有筛选
        for one_fitter in fitter_all_list:
            if self.fitter_list[0] == one_fitter[0] and self.fitter_list[1] == one_fitter[1]:
                if self.fitter_type == 'bybanner':  # banner有3个选项
                    if self.fitter_list[2] == one_fitter[2]:
                        fitter_id = one_fitter[-1]
                        break
                else:
                    fitter_id = one_fitter[-1]
                    break
        # print(table_fitter_pd)
        # print(fitter_id)
        fitter_df = table_fitter_pd[table_fitter_pd.tid.isin([fitter_id])].drop(['tid'],
                                                                                axis=1)  # 筛选 table_fitter_pd 里 tid列 中值为 fitter_id 的数据
        # fitter_df = table_fitter_pd[table_fitter_pd['tid']>20]                 #筛选 tabble_fitter_pd 里 tid列 中值大于 20 的数据
        # print(fitter_id)
        fitter_df = self.dealwith_data(fitter_df)
        # print(fitter_df)  # 下一步为插入对应表
        if self.fitter_type == 'byou' and 'Mini' in self.fitter_table:
            zz = fitter_df[[0, 1, 2, 22, 27, 28, 32]]  # ou表mini的选择列（特殊）
            # print(zz)
            all_data_list = []
            for one_data in zz.iterrows():
                # print(list(one_data[1]))
                all_data_list.append(list(one_data[1]))
            wb = openpyxl.load_workbook(
                self.file_title_path + '/{}/'.format(self.fitter_type) + self.fitter_table + '.src.xlsx')
            ws = wb.active
            mes_list = all_data_list[0]
            # print(mes_list)
            mes_zm = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
            self.table_index_list = mes_zm
            for j, one_mes in enumerate(all_data_list):  # 写单元格信息
                # print(one_mes)
                for z, one_sell in enumerate(one_mes):
                    ws['{}{}'.format(mes_zm[z], j + 3)] = one_sell
            wb.save('text.xlsx')
        else:
            all_data_list = []
            for one_data in fitter_df.iterrows():
                # print(list(one_data[1]))
                all_data_list.append(list(one_data[1]))
            wb = openpyxl.load_workbook(
                self.file_title_path + '/{}/'.format(self.fitter_type) + self.fitter_table + '.src.xlsx')
            ws = wb.active
            mes_list = all_data_list[0]

            if len(mes_list) <= 25:  # 处理标题的长度对应的字母
                mes_zm = zimu[1:len(mes_list)]
                # print(mes_zm)
            else:
                else_zm_len = len(mes_list) - 25
                mes_zm = zimu[1:]
                num = 0
                for f_zm in zimu:
                    if num >= else_zm_len:
                        break
                    for s_zm in zimu:
                        if num >= else_zm_len:
                            break
                        ex_zm = f_zm + s_zm
                        mes_zm.append(ex_zm)
                        num += 1
            # print(mes_zm)
            self.table_index_list = mes_zm
            for j, one_mes in enumerate(all_data_list):  # 写单元格信息
                # print(one_mes)
                for z, one_sell in enumerate(one_mes):
                    ws['{}{}'.format(mes_zm[z], j + 3)] = one_sell
            wb.save('text.xlsx')
        # print('生成测试表完成')
        end = time.time()
        print('根据对应的id找到对应的表 时间：%.2f秒' % (end - start))

    def get_index(self):
        """获取选择显示的编号"""
        # mes_zm = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
        #           'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']
        mes_zm = self.table_index_list
        index_line_list = []  # 目标索引
        second_lins_list = []
        first_lins_list = []
        if self.fitter_type == 'bybanner':
            if self.fitter_table == 'Total score.Hyper':
                select_list = ['Channel', 'Hyper']  # 选择的模块
                selectnd_list = ['Hyper']
                mes_zm.insert(0, 'A')
            elif self.fitter_table == 'Total score.Super':
                select_list = ['Channel', 'Super']  # 选择的模块
                selectnd_list = ['Super']
                mes_zm.insert(0, 'A')
            elif self.fitter_table == 'Total score.Mini':
                select_list = ['Channel', 'Mini']  # 选择的模块
                selectnd_list = ['Mini']
                mes_zm.insert(0, 'A')
            elif self.fitter_table == 'Availability.HS':
                select_list = ['Channel', 'Hyper', 'Super']  # 选择的模块
                selectnd_list = ['H/S', 'Hyper', 'Super']
                mes_zm.insert(0, 'A')
            elif self.fitter_table == 'Availability.Mini':
                select_list = ['Channel', 'Mini']  # 选择的模块
                selectnd_list = ['Mini']
                mes_zm.insert(0, 'A')
            else:  # 其他所欲偶都是全显示
                select_list = ['Channel', 'Hyper', 'Super', 'Mini']  # 选择的模块
                selectnd_list = ['H/S', 'Hyper', 'Super', 'Mini']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')
        else:
            if self.fitter_table == 'Total score.Mini':  # 除了mini模块，其他都是全展示
                select_list = ['SCCL']  # 选择的模块
                selectnd_list = ['']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')
            elif self.fitter_table == 'Availability.Mini':
                select_list = ['SCCL']  # 选择的模块
                selectnd_list = ['']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')
            else:
                select_list = ['BG', 'CBL', 'SCCL', 'ZH']  # 选择的模块
                selectnd_list = ['CBL', 'SCCL', 'ZH']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')

        # mes_zm.insert(0, 'A')
        wb = openpyxl.load_workbook("text.xlsx")
        ws = wb.active
        max_column = ws.max_column  # 最大列
        # print(max_column)
        rows = ws.rows
        count = 0
        start = 0  # 通过调整 start 的值进行选择第几行excel开始读取, 1 为第二行开始读取
        # 获取前两行的标题
        for j, row in enumerate(rows):
            if count < start:
                count += 1
                continue
            # print(j)
            if j == 0:
                for i in range(max_column):
                    # print(row[i].value)  # row[0].value 为当前行第一个单元格的值
                    first_lins_list.append(row[i].value)
            elif j == 1:
                for i in range(max_column):
                    # print(row[i].value)  # row[0].value 为当前行第一个单元格的值
                    second_lins_list.append(row[i].value)
            else:
                break
        con_flag = False
        if self.fitter_type == 'byou' and 'Mini' in self.fitter_table:
            index_line_list = ['C', 'E', 'F', 'G', 'H']
        else:
            for j, one_lins in enumerate(first_lins_list):
                if con_flag and one_lins is None:
                    index_line_list.append(mes_zm[j])
                else:
                    con_flag = False
                for one_select in select_list:
                    if one_select == one_lins:
                        index_line_list.append(mes_zm[j])
                        con_flag = True
                if j == len(first_lins_list) - 1 and one_lins:
                    index_line_list.append(mes_zm[j + 1])
            Channel_title = second_lins_list[:7]  # 切出Channel模块
            # print(Channel_title)
            if self.fitter_type == 'bybanner':
                for one_select in selectnd_list:  # 剔除选择的的模块（后边会剔除没选的模块）
                    if one_select in Channel_title:
                        Channel_title.remove(one_select)
                for j, one_lins in enumerate(second_lins_list):  # 剔除未选择的编号
                    for one_title in Channel_title:
                        if one_title == one_lins and one_title:
                            index_line_list.remove(mes_zm[j])
            else:
                pass
        # print(index_line_list)
        return index_line_list

    def get_complete_table(self):
        """获取并写出无样式的表"""
        start = time.time()
        index = self.get_index()
        index = ['A'] + index
        # print(index)
        wb = openpyxl.load_workbook('text.xlsx')
        ws = wb.active
        name_fit_list = [x.replace('/', '') for x in self.fitter_list]  # 处理名字特殊符号
        if not os.path.exists(
                self.download_path + '{} {}.xlsx'.format(self.fitter_type, self.fitter_table)):  # 如果文件不存在
            wb_ = openpyxl.Workbook()
            wb_.create_sheet('_'.join(name_fit_list))  # 新建一个筛选项的sheet
            ws_ = wb_['_'.join(name_fit_list)]
        else:  # 存在就直接打开
            wb_ = openpyxl.load_workbook(
                self.download_path + '{} {}.xlsx'.format(self.fitter_type, self.fitter_table))
            sheet_name = wb_.sheetnames
            if '_'.join(name_fit_list) in sheet_name:  # 判断是否生成过，生成过就直接跳过
                return 'had_down'
            wb_.create_sheet('_'.join(name_fit_list))  # 新建一个筛选项的sheet
            ws_ = wb_['_'.join(name_fit_list)]
        all_row = ws.max_row  # 获取最大行数
        print('开始')
        for one_row in range(1, all_row + 1):
            lins_list = []
            for one_index in index:
                # print(one_index, one_row)
                one_value = ws['{}{}'.format(one_index, one_row)].value
                if one_value is None:
                    one_value = ''
                lins_list.append(one_value)
                # print(ws['{}{}'.format(one_index, one_row)].value)
            # print(lins_list)
            ws_.append(lins_list)
        wb_.save(self.download_path + '{} {}.xlsx'.format(self.fitter_type, self.fitter_table))
        # print('写表完成')
        end = time.time()
        print('获取并写出无样式的表 时间：%.2f秒' % (end - start))

    def dealwith_data(self, fitter_df):
        """处理各类不同的格式问题"""
        # 处理保留多少位数或者百分号
        if self.fitter_type == 'bybanner':
            if self.fitter_table == 'Total score.Hyper' or self.fitter_table == 'Total score.Super' or self.fitter_table == 'Total score.Mini':  # 直接保留 1 位小数
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        fitter_df[one_title] = fitter_df[one_title].round(decimals=1)
                    except Exception as e:
                        print(e)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif self.fitter_table == 'Availability.HS' or self.fitter_table == 'Availability.Mini' or self.fitter_table == 'Thematic Display':  # 全部百分号
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        # fitter_df[one_title] = fitter_df[one_title].round(decimals=1)
                        print(fitter_df[one_title][2:])
                        fitter_df[one_title][2:] = fitter_df[one_title][2:].apply(lambda x: format(x, '.0%'))
                    except Exception as e:
                        print(e)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
                # 清除空行
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        try:
                            if data == '0':
                                fitter_df.iloc[j, 1:] = ''
                        except Exception as e:
                            print(e)
            elif self.fitter_table == '2nd Display':  # 处理：1、清除空行数据 2、改部分数据为%  3、有'NET' 数据为空
                for one_row in fitter_df.iterrows():
                    index = one_row[0]
                    if index != 0:
                        data = one_row[1][0]
                        try:
                            if data != '0' and '（m2)' not in data:
                                # print(fitter_df.iloc[index, :])
                                if 'NET' in data:
                                    fitter_df.iloc[index, 1:] = ''
                                else:
                                    fitter_df.iloc[index, 1:] = fitter_df.iloc[index, 1:].apply(
                                        lambda x: format(x, '.0%'))
                            if data == '0':
                                fitter_df.iloc[index, 1:] = ''
                                # print(data)
                        except Exception as e:
                            print(e)
                            pass
                        # print(data)
            elif self.fitter_table == 'Cooler':  # 处理：1、纯净度 改为 % 2、清除空列
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        try:
                            if '纯净度' in data or '全品类' in data or '汽水' in data or '果汁' in data or '包装水' in data:  # 百分号
                                # print(fitter_df.iloc[index, :])
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                            elif data == '0' or '单位' in data:  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            else:  # 保留2位数
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: '%.2f' % x)
                        except Exception as e:
                            print(e)
                            pass
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif self.fitter_table == 'SOVI':
                percentage_count = -1
                start_flag = False
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        if start_flag:  # 计数开始
                            if data == '0':  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            if percentage_count == 3:  # 百分比
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                                percentage_count = 0
                            elif percentage_count == 2 or percentage_count == 1:  # 保留2位数
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: '%.2f' % x)
                            percentage_count += 1
                        else:
                            if data == '0':  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            else:
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                        if data == 'KO Plant Protein & VAD SOVI':
                            start_flag = True
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif self.fitter_table == 'Price':
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        if data == '0':  # 空行
                            fitter_df.iloc[j, 1:] = ''
                        elif '占比' in data:
                            fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                lambda x: format(x, '.0%'))
                        else:
                            fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                lambda x: '%.2f' % x)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
        else:  # byou 的
            try:
                if self.fitter_table == 'Total score.Hyper' or self.fitter_table == 'Total score.Super' or self.fitter_table == 'Total score.Mini':  # 直接保留 1 位小数
                    for one_title in list(fitter_df.columns.values)[1:]:
                        try:
                            fitter_df[one_title] = fitter_df[one_title].round(decimals=1)
                        except Exception as e:
                            print(e)
                    # 清除空列
                    for one_title in list(fitter_df.columns.values)[1:]:
                        try:
                            if fitter_df.iloc[0, one_title]:  # 如果门店
                                pass
                            else:
                                fitter_df[one_title] = ''
                        except Exception as e:
                            print(e)
                    # 清除空行
                    for j, one_row in enumerate(fitter_df.iterrows()):
                        if j != 0:
                            data = one_row[1][0]
                            try:
                                if data == '0':
                                    fitter_df.iloc[j, 1:] = ''
                            except Exception as e:
                                print(e)
                elif self.fitter_table == 'Availability.HS' or self.fitter_table == 'Availability.Mini' or self.fitter_table == 'Thematic Display':  # 全部百分号
                    for one_title in list(fitter_df.columns.values)[1:]:
                        try:
                            fitter_df.loc[1:, one_title] = fitter_df.loc[1:, one_title].round(decimals=2).apply(
                                lambda x: format(x, '.0%'))
                        except Exception as e:
                            print(e)
                    # 清除空列
                    for one_title in list(fitter_df.columns.values)[1:]:
                        try:
                            if fitter_df.iloc[0, one_title]:  # 如果门店
                                pass
                            else:
                                fitter_df[one_title] = ''
                        except Exception as e:
                            print(e)
                    # 清除空行
                    for j, one_row in enumerate(fitter_df.iterrows()):
                        if j != 0 and j != 2:  # 第二行也不可清除
                            data = one_row[1][0]
                            try:
                                if data == '0':
                                    fitter_df.iloc[j, 1:] = ''
                            except Exception as e:
                                print(e)
                elif self.fitter_table == '2nd Display':  # 处理：1、清除空行数据 2、改部分数据为%  3、有'NET' 数据为空
                    for j, one_row in enumerate(fitter_df.iterrows()):
                        if j != 0:
                            data = one_row[1][0]
                            try:
                                if data != '0' and '（m2)' not in data and j != 2:
                                    # print(fitter_df.iloc[index, :])
                                    if 'NET' in data:
                                        fitter_df.iloc[j, 1:] = ''
                                    else:
                                        fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                            lambda x: format(x, '.0%'))
                                elif data == '0' and j != 2:
                                    fitter_df.iloc[j, 1:] = ''
                                elif j == 2:  # j == 2 的情况
                                    fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                        lambda x: format(x, '.0%'))
                            except Exception as e:
                                print(e)
                                pass
                elif self.fitter_table == 'Cooler':  # 处理：1、纯净度 改为 % 2、清除空列
                    for j, one_row in enumerate(fitter_df.iterrows()):
                        if j != 0:
                            data = one_row[1][0]
                            try:
                                if '纯净度' in data or '全品类' in data or '汽水' in data or '果汁' in data or '包装水' in data:  # 百分号
                                    # print(fitter_df.iloc[index, :])
                                    fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                        lambda x: format(x, '.0%'))
                                elif data == '0' or '单位' in data:  # 空行
                                    fitter_df.iloc[j, 1:] = ''
                                else:  # 保留2位数
                                    fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                        lambda x: '%.2f' % x)
                            except Exception as e:
                                print(e)
                                pass
                    # 清除空列
                    for one_title in list(fitter_df.columns.values)[1:]:
                        try:
                            if fitter_df.iloc[0, one_title]:  # 如果门店
                                pass
                            else:
                                fitter_df[one_title] = ''
                        except Exception as e:
                            print(e)
                elif self.fitter_table == 'SOVI':
                    percentage_count = -1
                    start_flag = False
                    for j, one_row in enumerate(fitter_df.iterrows()):
                        if j != 0:
                            data = one_row[1][0]
                            if start_flag:  # 计数开始
                                if data == '0':  # 空行
                                    fitter_df.iloc[j, 1:] = ''
                                if percentage_count == 3:  # 百分比
                                    fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                        lambda x: format(x, '.0%'))
                                    percentage_count = 0
                                elif percentage_count == 2 or percentage_count == 1:  # 保留2位数
                                    fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                        lambda x: '%.2f' % x)
                                percentage_count += 1
                            else:
                                if data == '0':  # 空行
                                    fitter_df.iloc[j, 1:] = ''
                                else:
                                    fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                        lambda x: format(x, '.0%'))
                            if data == 'KO Plant Protein & VAD SOVI':
                                start_flag = True
                    # 清除空列
                    for one_title in list(fitter_df.columns.values)[1:]:
                        try:
                            if fitter_df.iloc[0, one_title]:  # 如果门店
                                pass
                            else:
                                fitter_df[one_title] = ''
                        except Exception as e:
                            print(e)
                elif self.fitter_table == 'Price':
                    for j, one_row in enumerate(fitter_df.iterrows()):
                        if j != 0:
                            data = one_row[1][0]
                            if data == '0':  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            elif '占比' in data:
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                            else:
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: '%.2f' % x)
                    # 清除空列
                    for one_title in list(fitter_df.columns.values)[1:]:
                        try:
                            if fitter_df.iloc[0, one_title]:  # 如果门店
                                pass
                            else:
                                fitter_df[one_title] = ''
                        except Exception as e:
                            print(e)
            except Exception as e:
                print(e)
        # print(fitter_df)
        # 清除空行
        return fitter_df

    def wirte_style(self):
        """写样式"""
        start = time.time()
        zimu = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
        wb = openpyxl.load_workbook(self.download_path + '{} {}.xlsx'.format(self.fitter_type, self.fitter_table))
        name_fit_list = [x.replace('/', '') for x in self.fitter_list]  # 处理名字特殊符号
        ws = wb['_'.join(name_fit_list)]
        z = ws.max_column  # 最大列
        if z <= 26:  # 处理标题的长度对应的字母
            mes_zm = zimu[:z]
            # print(mes_zm)
        else:
            else_zm_len = z - 26
            mes_zm = zimu
            num = 0
            for f_zm in zimu:
                if num >= else_zm_len:
                    break
                for s_zm in zimu:
                    if num >= else_zm_len:
                        break
                    ex_zm = f_zm + s_zm
                    mes_zm.append(ex_zm)
                    num += 1
        Channal_index_list = []  # Channal 的标题
        # else_index_lise = []  # 其他的标题
        if self.fitter_type == 'bybanner':
            all_Channal = ['H/S', 'Hyper', 'Super', 'Mini']
        else:
            all_Channal = ['CBL', 'SCCL', 'ZH']

        title_font = Font(name='Arial', size=10, bold=True)  # 字体
        font = Font(name='Arial', size=10)  # 字体
        align = Alignment(horizontal='center')  # 居中
        fill = PatternFill("solid", fgColor='E5E5E5')  # 标题背景颜色
        shop_fill = PatternFill("solid", fgColor='C1FFC1')  # 标题背景颜色
        color_list = ['27408B', 'FF0000', '4876FF', '63B8FF']  # 其他字体颜色
        # else_title = mes_zm[mes_zm.index(previous_) + 1]  # 获取其他标题
        if self.fitter_type == 'byou' and 'Mini' in self.fitter_table:  # Mini 的只合并这一种
            ws.merge_cells('C1:F1')
        else:
            # 读第二行 判断合并位置
            for one_index in mes_zm[:7]:
                if ws['{}2'.format(one_index)].value in all_Channal:
                    Channal_index_list.append(one_index)
            # 判断其他的标题的开始位置
            previous_ = Channal_index_list[-1]
            else_index = mes_zm[mes_zm.index(previous_) + 1:]
            # else_index = zimu
            else_title = [ws['{}1'.format(x)].value for x in else_index]
            ws.merge_cells('{}1:{}1'.format(Channal_index_list[0], Channal_index_list[-1]))  # 合并Channal
            merge_flag = False  # 合并标志
            for j, one_else_title in enumerate(else_title):
                if one_else_title:
                    if merge_flag == False:
                        lins_list = []
                        lins_list.append(else_index[j])
                        merge_flag = True
                    else:
                        ws.merge_cells('{}1:{}1'.format(lins_list[0], lins_list[-1]))  # 合并其他单元格
                        merge_flag = False
                        lins_list = []
                        lins_list.append(else_index[j])
                        merge_flag = True
                else:
                    lins_list.append(else_index[j])
                if j == len(else_title) - 1:
                    ws.merge_cells('{}1:{}1'.format(lins_list[0], lins_list[-1]))  # 合并其他单元格
        if self.fitter_type == 'bybanner':
            ws['{}1'.format(Channal_index_list[0])] = "Channal"
        else:
            pass
        # 标题格式：居中，加粗,底色
        for one_size in mes_zm:
            if ws['{}1'.format(one_size)].value == None:
                continue
            else:
                ws['{}1'.format(one_size)].font = title_font
                ws['{}1'.format(one_size)].alignment = align
                ws['{}1'.format(one_size)].fill = fill
        # 门店数量格式：字体居中，底色
        for one_size in mes_zm:
            ws['{}3'.format(one_size)].font = font
            ws['{}3'.format(one_size)].alignment = align
            ws['{}3'.format(one_size)].fill = shop_fill
        # 其他位置格式：字体居中
        for i in range(2, ws.max_row + 1):
            if '满分' in str(ws['A{}'.format(i)].value):  # 特殊行数要颜色
                if i == 5:
                    font = Font(name='Arial', size=10, color=color_list[0], bold=True)
                elif i == 6:
                    font = Font(name='Arial', size=10, color=color_list[1], bold=True)
                elif i == 7:
                    font = Font(name='Arial', size=10, color=color_list[2], bold=True)
                elif i == 8:
                    font = Font(name='Arial', size=10, color=color_list[3], bold=True)
                else:
                    font = Font(name='Arial', size=10, color=color_list[2], bold=True)
            else:
                font = Font(name='Arial', size=10)  # 普通字体
            for one_size in mes_zm:
                ws['{}{}'.format(one_size, i)].font = font
                ws['{}{}'.format(one_size, i)].alignment = align
        wb.save(self.download_path + '{} {}.xlsx'.format(self.fitter_type, self.fitter_table))
        # print('样式完成')
        end = time.time()
        print('写样式 时间：%.2f秒' % (end - start))

    def save_transpose(self):
        wb = openpyxl.load_workbook(self.store_path + '{}.xlsx'.format(self.fitter_table))
        wb.save(self.download_path + '{} {}.xlsx'.format(self.fitter_type, self.fitter_table))

    def run(self):
        # if self.fitter_type == 'bybanner':
        #     self.get_banner_fitter_id() #获取id
        # else:
        #     self.get_ou_fitter_id()  # 获取id
        start = time.time()
        if 'by Store' in self.fitter_table:
            self.save_transpose()
            self.str_out.emit('{} {}.xlsx 生成选项表成功'.format(self.fitter_type, self.fitter_table), self.fitter_type)
        else:
            self.get_fitter_table()  # 生成基础表
            flag = self.get_complete_table()  # 生成无样式的完整表
            if flag != 'had_down':  # 未生成过
                self.wirte_style()  # 写样式
                print('完成')
                self.str_out.emit('{} {}.xlsx 生成选项表成功'.format(self.fitter_type, self.fitter_table), self.fitter_type)
            else:
                self.str_out.emit('{} {}.xlsx 已生成过此选项'.format(self.fitter_type, self.fitter_table), self.fitter_type)
        end = time.time()
        self.x = print(end - start)


class GenerateTableShow(GenerateTable):
    """生成完整的table(展示)"""
    str_out = pyqtSignal(str, str)  # 打印窗口信号
    clear_str = pyqtSignal(str) #清楚打印窗口
    title_out = pyqtSignal(list, str)  # 标题发送信号
    data_out = pyqtSignal(list, str)  # 数据信号
    clear_out = pyqtSignal(str)  # 清除数据

    def __init__(self, fitter_type, fitter_table, fitter_list):
        super().__init__(fitter_type, fitter_table, fitter_list)
        # self.setting_file = '测试设置表.xlsx'
        # self.bannerS_file = 'bybannerSS.xlsx'
        # self.ouS_file = 'byouSS.xlsx'
        # self.file_title_path = './stc_title/'  # 表头路径
        # self.fitter_type = fitter_type  # 表类型
        # self.fitter_list = fitter_list  # 对应的id 为7
        # self.fitter_table = fitter_table  # 表分类
        # self.table_index_list = []  # 获取到的表的id

    def get_complete_table(self):
        """获取并写出无样式的表"""
        start = time.time()
        index = self.get_index()
        index = ['A'] + index
        print(index)
        wb = openpyxl.load_workbook('text.xlsx')
        ws = wb.active
        # wb_ = openpyxl.load_workbook('banner_complete.xlsx')
        # wb_ = openpyxl.Workbook()
        # # wb.create_sheet('TableCopy')  # 新建一个sheet
        # ws_ = wb_.active
        all_row = ws.max_row  # 获取最大行数
        print('开始')
        all_data_list = []
        for one_row in range(1, all_row + 1):
            lins_list = []
            for one_index in index:
                # print(one_index, one_row)
                one_value = ws['{}{}'.format(one_index, one_row)].value
                if one_value is None:
                    one_value = ''
                lins_list.append(one_value)
                # print(ws['{}{}'.format(one_index, one_row)].value)
            # print(lins_list)
            all_data_list.append(lins_list)
        # wb_.save('{}_complete.xlsx'.format(self.fitter_type))
        # print('写表完成')
        end = time.time()
        print('获取并写出无样式的表 时间：%.2f秒' % (end - start))
        return all_data_list

    def get_data_title(self, title_list):
        """获取标题并传过去"""
        if self.fitter_type == 'bybanner':
            first_title = title_list[0]
            second_title = title_list[1]
            all_title = []
            first_t = ''
            for j, one_second in enumerate(second_title):
                if j == 0:
                    all_title.append('')
                    continue
                if j == 1 and first_title[j] == '':  # 添加第一块标题
                    first_t = 'Channal'
                if first_title[j]:  # 如果有就重新更新第一个标题
                    first_t = first_title[j]
                if first_title[j] == '' and one_second == '':  # 两个都没就加一行空
                    all_title.append('')
                else:
                    all_title.append(first_t + '/' + one_second)
            print(all_title)
            self.title_out.emit(all_title, self.fitter_type)
        else:
            first_title = title_list[0]
            second_title = title_list[1]
            all_title = []
            first_t = ''
            for j, one_second in enumerate(second_title):
                if j == 0:
                    all_title.append('')
                    continue
                if j == 1 and first_title[j] == '':  # 添加第一块标题
                    first_t = 'Channal'
                if first_title[j]:  # 如果有就重新更新第一个标题
                    first_t = first_title[j]
                if first_title[j] == '' and one_second == '':  # 两个都没就加一行空
                    all_title.append('')
                else:
                    all_title.append(first_t + '/' + one_second)
            # print(all_title)
            self.title_out.emit(all_title, self.fitter_type)

    def get_data(self, data):
        for one_data in data:
            self.data_out.emit(one_data, self.fitter_type)

    # 转置操作
    def get_transpose_data(self):
        wb = openpyxl.load_workbook(self.store_path + '{}.xlsx'.format(self.fitter_table))
        ws = wb.active
        all_row = ws.max_column  # 获取最大行数
        # print('开始')
        # print(all_row)
        all_data_list = []
        rows = ws.rows
        count = 0
        if self.fitter_table == 'SOVI by Store':
            start = 2  # 通过调整 start 的值进行选择第几行excel开始读取, 1 为第二行开始读取
        else:
            start = 1
        for j, row in enumerate(rows):
            if count < start:
                count += 1
                continue
            lins_list = []
            for i in range(all_row):
                lins_list.append(row[i].value)
            # print(lins_list)
            all_data_list.append(lins_list)
        return all_data_list

    def set_transpose_title(self, title_list):
        # print('表头')
        first_title = title_list[0]
        second_title = title_list[1]
        all_title = []
        for j, one_title in enumerate(first_title):
            if one_title and second_title[j]:
                all_title.append(one_title + '/' + second_title[j])
            elif one_title and second_title[j] == None:
                all_title.append(one_title)
            elif one_title == None and second_title[j]:
                all_title.append(second_title[j])
            else:
                all_title.append('')
        self.title_out.emit(all_title, self.fitter_type)

    def set_transpose_data(self, data):
        for one_data in data:
            self.data_out.emit(one_data, self.fitter_type)

    def run(self):
        if 'by Store' in self.fitter_table:  # by Store 表
            all_data = self.get_transpose_data()
            self.set_transpose_title(all_data[:2])  # 设置表头
            self.set_transpose_data(all_data[2:])  # 设置数据
        else:
            self.get_fitter_table()  # 生成基础表
            all_data = self.get_complete_table()  # 生成无样式的完整表
            self.clear_out.emit(self.fitter_type)  # 清屏
            self.get_data_title((all_data[:2]))
            self.get_data(all_data[2:])
        self.clear_str.emit(self.fitter_type)
        self.str_out.emit('查询完毕{}_{}'.format(self.fitter_table,'_'.join(self.fitter_list)), self.fitter_type)
        # print(self.fitter_list)


class DownloadAllFiles(GenerateTable):
    """下载所有表"""
    str_out = pyqtSignal(str, str)  # 打印窗口信号
    count_out = pyqtSignal(int, str)  # 统计打印

    def __init__(self, fitter_type, fitter_table, fitter_list, banner_table_dict, ou_table_dict):
        super().__init__(fitter_type, fitter_table, fitter_list)
        self.banner_table_dict = banner_table_dict  # banner 所有选项
        self.ou_table_dict = ou_table_dict  # uo 所有选项
        self.lins_file_path = './lins_file/'
        if not os.path.exists(self.lins_file_path):  # 新建文件夹
            os.makedirs(self.lins_file_path)
        self.complete_num = 0
        self.all_start = 0
        self.all_end = 0

    def dealwith_data(self, fitter_df, now_fitter_type=None, now_fitter_table=None):
        """处理各类不同的格式问题"""
        # 处理保留多少位数或者百分号
        if now_fitter_type == 'bybanner':
            if now_fitter_table == 'Total score.Hyper' or now_fitter_table == 'Total score.Super' or now_fitter_table == 'Total score.Mini':  # 直接保留 1 位小数
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        fitter_df[one_title] = fitter_df[one_title].round(decimals=1)
                    except Exception as e:
                        print(e)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif now_fitter_table == 'Availability.HS' or now_fitter_table == 'Availability.Mini' or now_fitter_table == 'Thematic Display':  # 全部百分号
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        # fitter_df[one_title] = fitter_df[one_title].round(decimals=1)
                        # print(fitter_df[one_title][2:])
                        fitter_df[one_title][2:] = fitter_df[one_title][2:].apply(lambda x: format(x, '.0%'))
                    except Exception as e:
                        print(e)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
                # 清除空行
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        try:
                            if data == '0':
                                fitter_df.iloc[j, 1:] = ''
                        except Exception as e:
                            print(e)
            elif now_fitter_table == '2nd Display':  # 处理：1、清除空行数据 2、改部分数据为%  3、有'NET' 数据为空
                for one_row in fitter_df.iterrows():
                    index = one_row[0]
                    if index != 0:
                        data = one_row[1][0]
                        try:
                            if data != '0' and '（m2)' not in data:
                                # print(fitter_df.iloc[index, :])
                                if 'NET' in data:
                                    fitter_df.iloc[index, 1:] = ''
                                else:
                                    fitter_df.iloc[index, 1:] = fitter_df.iloc[index, 1:].apply(
                                        lambda x: format(x, '.0%'))
                            if data == '0':
                                fitter_df.iloc[index, 1:] = ''
                                # print(data)
                        except Exception as e:
                            print(e)
                            pass
                        # print(data)
            elif now_fitter_table == 'Cooler':  # 处理：1、纯净度 改为 % 2、清除空列
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        try:
                            if '纯净度' in data or '全品类' in data or '汽水' in data or '果汁' in data or '包装水' in data:  # 百分号
                                # print(fitter_df.iloc[index, :])
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                            elif data == '0' or '单位' in data:  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            else:  # 保留2位数
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: '%.2f' % x)
                        except Exception as e:
                            print(e)
                            pass
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif now_fitter_table == 'SOVI':
                percentage_count = -1
                start_flag = False
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        if start_flag:  # 计数开始
                            if data == '0':  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            if percentage_count == 3:  # 百分比
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                                percentage_count = 0
                            elif percentage_count == 2 or percentage_count == 1:  # 保留2位数
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: '%.2f' % x)
                            percentage_count += 1
                        else:
                            if data == '0':  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            else:
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                        if data == 'KO Plant Protein & VAD SOVI':
                            start_flag = True
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif now_fitter_table == 'Price':
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        if data == '0':  # 空行
                            fitter_df.iloc[j, 1:] = ''
                        elif '占比' in data:
                            fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                lambda x: format(x, '.0%'))
                        else:
                            fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                lambda x: '%.2f' % x)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
        else:  # byou 的
            if now_fitter_table == 'Total score.Hyper' or now_fitter_table == 'Total score.Super' or now_fitter_table == 'Total score.Mini':  # 直接保留 1 位小数
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        fitter_df[one_title] = fitter_df[one_title].round(decimals=1)
                    except Exception as e:
                        print(e)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
                # 清除空行
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        try:
                            if data == '0':
                                fitter_df.iloc[j, 1:] = ''
                        except Exception as e:
                            print(e)
            elif now_fitter_table == 'Availability.HS' or now_fitter_table == 'Availability.Mini' or now_fitter_table == 'Thematic Display':  # 全部百分号
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        fitter_df.loc[1:, one_title] = fitter_df.loc[1:, one_title].round(decimals=2).apply(
                            lambda x: format(x, '.0%'))
                    except Exception as e:
                        print(e)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
                # 清除空行
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0 and j != 2:  # 第二行也不可清除
                        data = one_row[1][0]
                        try:
                            if data == '0':
                                fitter_df.iloc[j, 1:] = ''
                        except Exception as e:
                            print(e)
            elif now_fitter_table == '2nd Display':  # 处理：1、清除空行数据 2、改部分数据为%  3、有'NET' 数据为空
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        try:
                            if data != '0' and '（m2)' not in data and j != 2:
                                # print(fitter_df.iloc[index, :])
                                if 'NET' in data:
                                    fitter_df.iloc[j, 1:] = ''
                                else:
                                    fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                        lambda x: format(x, '.0%'))
                            elif data == '0' and j != 2:
                                fitter_df.iloc[j, 1:] = ''
                            elif j == 2:  # j == 2 的情况
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                        except Exception as e:
                            print(e)
                            pass
            elif now_fitter_table == 'Cooler':  # 处理：1、纯净度 改为 % 2、清除空列
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        try:
                            if '纯净度' in data or '全品类' in data or '汽水' in data or '果汁' in data or '包装水' in data:  # 百分号
                                # print(fitter_df.iloc[index, :])
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                            elif data == '0' or '单位' in data:  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            else:  # 保留2位数
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: '%.2f' % x)
                        except Exception as e:
                            print(e)
                            pass
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif now_fitter_table == 'SOVI':
                percentage_count = -1
                start_flag = False
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        if start_flag:  # 计数开始
                            if data == '0':  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            if percentage_count == 3:  # 百分比
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                                percentage_count = 0
                            elif percentage_count == 2 or percentage_count == 1:  # 保留2位数
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: '%.2f' % x)
                            percentage_count += 1
                        else:
                            if data == '0':  # 空行
                                fitter_df.iloc[j, 1:] = ''
                            else:
                                fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                    lambda x: format(x, '.0%'))
                        if data == 'KO Plant Protein & VAD SOVI':
                            start_flag = True
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
            elif now_fitter_table == 'Price':
                for j, one_row in enumerate(fitter_df.iterrows()):
                    if j != 0:
                        data = one_row[1][0]
                        if data == '0':  # 空行
                            fitter_df.iloc[j, 1:] = ''
                        elif '占比' in data:
                            fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                lambda x: format(x, '.0%'))
                        else:
                            fitter_df.iloc[j, 1:] = fitter_df.iloc[j, 1:].apply(
                                lambda x: '%.2f' % x)
                # 清除空列
                for one_title in list(fitter_df.columns.values)[1:]:
                    try:
                        if fitter_df.iloc[0, one_title]:  # 如果门店
                            pass
                        else:
                            fitter_df[one_title] = ''
                    except Exception as e:
                        print(e)
        # print(fitter_df)
        # 清除空行
        return fitter_df

    def get_index(self, now_fitter_type=None, now_fitter_table=None, now_index_list=None):
        """获取选择显示的编号"""
        # mes_zm = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
        #           'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']
        mes_zm = now_index_list
        index_line_list = []  # 目标索引
        second_lins_list = []
        first_lins_list = []
        if now_fitter_type == 'bybanner':
            if now_fitter_table == 'Total score.Hyper':
                select_list = ['Channel', 'Hyper']  # 选择的模块
                selectnd_list = ['Hyper']
                mes_zm.insert(0, 'A')
            elif now_fitter_table == 'Total score.Super':
                select_list = ['Channel', 'Super']  # 选择的模块
                selectnd_list = ['Super']
                mes_zm.insert(0, 'A')
            elif now_fitter_table == 'Total score.Mini':
                select_list = ['Channel', 'Mini']  # 选择的模块
                selectnd_list = ['Mini']
                mes_zm.insert(0, 'A')
            elif now_fitter_table == 'Availability.HS':
                select_list = ['Channel', 'Hyper', 'Super']  # 选择的模块
                selectnd_list = ['H/S', 'Hyper', 'Super']
                mes_zm.insert(0, 'A')
            elif now_fitter_table == 'Availability.Mini':
                select_list = ['Channel', 'Mini']  # 选择的模块
                selectnd_list = ['Mini']
                mes_zm.insert(0, 'A')
            else:  # 其他所欲偶都是全显示
                select_list = ['Channel', 'Hyper', 'Super', 'Mini']  # 选择的模块
                selectnd_list = ['H/S', 'Hyper', 'Super', 'Mini']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')
        else:
            if now_fitter_table == 'Total score.Mini':  # 除了mini模块，其他都是全展示
                select_list = ['SCCL']  # 选择的模块
                selectnd_list = ['']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')
            elif now_fitter_table == 'Availability.Mini':
                select_list = ['SCCL']  # 选择的模块
                selectnd_list = ['']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')
            else:
                select_list = ['BG', 'CBL', 'SCCL', 'ZH']  # 选择的模块
                selectnd_list = ['CBL', 'SCCL', 'ZH']
                mes_zm.insert(0, 'B')
                mes_zm.insert(0, 'A')
        # mes_zm.insert(0, 'A')
        wb = openpyxl.load_workbook(self.lins_file_path + "text_{}.xlsx".format(now_fitter_table))
        ws = wb.active
        max_column = ws.max_column  # 最大列
        # print(max_column)
        rows = ws.rows
        count = 0
        start = 0  # 通过调整 start 的值进行选择第几行excel开始读取, 1 为第二行开始读取
        # 获取前两行的标题
        for j, row in enumerate(rows):
            if count < start:
                count += 1
                continue
            # print(j)
            if j == 0:
                for i in range(max_column):
                    # print(row[i].value)  # row[0].value 为当前行第一个单元格的值
                    first_lins_list.append(row[i].value)
            elif j == 1:
                for i in range(max_column):
                    # print(row[i].value)  # row[0].value 为当前行第一个单元格的值
                    second_lins_list.append(row[i].value)
            else:
                break
        con_flag = False
        if now_fitter_type == 'byou' and 'Mini' in now_fitter_table:
            index_line_list = ['C', 'E', 'F', 'G', 'H']
        else:
            for j, one_lins in enumerate(first_lins_list):
                if con_flag and one_lins is None:
                    index_line_list.append(mes_zm[j])
                else:
                    con_flag = False
                for one_select in select_list:
                    if one_select == one_lins:
                        index_line_list.append(mes_zm[j])
                        con_flag = True
                if j == len(first_lins_list) - 1 and one_lins:
                    index_line_list.append(mes_zm[j + 1])
            Channel_title = second_lins_list[:7]  # 切出Channel模块
            # print(Channel_title)
            if now_fitter_type == 'bybanner':
                for one_select in selectnd_list:  # 剔除选择的的模块（后边会剔除没选的模块）
                    if one_select in Channel_title:
                        Channel_title.remove(one_select)
                for j, one_lins in enumerate(second_lins_list):  # 剔除未选择的编号
                    for one_title in Channel_title:
                        if one_title == one_lins and one_title:
                            index_line_list.remove(mes_zm[j])
            else:
                pass
        # print(index_line_list)
        return index_line_list

    def get_fitter_table(self, now_fitter_type=None, now_fitter_table=None, now_fitter_list=None):
        """根据对应的id找到对应的表"""
        zimu = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
        fitter_id = ''
        # start = time.time()
        if now_fitter_type == 'bybanner':
            table_fitter_pd = all_banner_data[now_fitter_table]
            fitter_all_list = self.get_banner_fitter_id()  # 获取所有筛选
        else:  # ou表
            # table_fitter_pd = pd.read_excel(self.ouS_file, sheet_name=now_fitter_table + '.src')
            table_fitter_pd = all_ou_data[now_fitter_table]
            fitter_all_list = self.get_ou_fitter_id()  # 获取所有筛选
        for one_fitter in fitter_all_list:
            if now_fitter_list[0] == one_fitter[0] and now_fitter_list[1] == one_fitter[1]:
                if now_fitter_type == 'bybanner':  # banner有3个选项
                    if now_fitter_list[2] == one_fitter[2]:
                        fitter_id = one_fitter[-1]
                        break
                else:
                    fitter_id = one_fitter[-1]
                    break
        # print(table_fitter_pd)
        fitter_df = table_fitter_pd[table_fitter_pd.tid.isin([fitter_id])].drop(['tid'],
                                                                                axis=1)  # 筛选 table_fitter_pd 里 tid列 中值为 fitter_id 的数据
        # fitter_df = table_fitter_pd[table_fitter_pd['tid']>20]                 #筛选 tabble_fitter_pd 里 tid列 中值大于 20 的数据
        # print(fitter_id)
        fitter_df = self.dealwith_data(fitter_df, now_fitter_type, now_fitter_table)
        # print(fitter_df)  # 下一步为插入对应表
        if now_fitter_type == 'byou' and 'Mini' in now_fitter_table:
            zz = fitter_df[[0, 1, 2, 22, 27, 28, 32]]  # ou表mini的选择列（特殊）
            # print(zz)
            all_data_list = []
            for one_data in zz.iterrows():
                # print(list(one_data[1]))
                all_data_list.append(list(one_data[1]))
            wb = openpyxl.load_workbook(
                self.file_title_path + '/{}/'.format(now_fitter_type) + now_fitter_table + '.src.xlsx')
            ws = wb.active
            mes_list = all_data_list[0]
            # print(mes_list)
            mes_zm = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
            # self.table_index_list = mes_zm
            for j, one_mes in enumerate(all_data_list):  # 写单元格信息
                # print(one_mes)
                for z, one_sell in enumerate(one_mes):
                    ws['{}{}'.format(mes_zm[z], j + 3)] = one_sell
            wb.save(self.lins_file_path + "text_{}.xlsx".format(now_fitter_table))
        else:
            all_data_list = []
            for one_data in fitter_df.iterrows():
                # print(list(one_data[1]))
                all_data_list.append(list(one_data[1]))
            wb = openpyxl.load_workbook(
                self.file_title_path + '/{}/'.format(now_fitter_type) + now_fitter_table + '.src.xlsx')
            ws = wb.active
            mes_list = all_data_list[0]

            if len(mes_list) <= 25:  # 处理标题的长度对应的字母
                mes_zm = zimu[1:len(mes_list)]
                # print(mes_zm)
            else:
                else_zm_len = len(mes_list) - 25
                mes_zm = zimu[1:]
                num = 0
                for f_zm in zimu:
                    if num >= else_zm_len:
                        break
                    for s_zm in zimu:
                        if num >= else_zm_len:
                            break
                        ex_zm = f_zm + s_zm
                        mes_zm.append(ex_zm)
                        num += 1
            # print(mes_zm)
            # self.table_index_list = mes_zm
            for j, one_mes in enumerate(all_data_list):  # 写单元格信息
                # print(one_mes)
                for z, one_sell in enumerate(one_mes):
                    ws['{}{}'.format(mes_zm[z], j + 3)] = one_sell
            wb.save(self.lins_file_path + "text_{}.xlsx".format(now_fitter_table))
        # print('生成测试表完成')
        # end = time.time()
        # print('根据对应的id找到对应的表 时间：%.2f秒' % (end - start))
        return mes_zm

    def get_complete_table(self, now_fitter_type=None, now_fitter_table=None, now_fitter_list=None, now_mes_zm=None):
        """
        生成完整无样式表
        :return:
        """
        # start = time.time()
        index = self.get_index(now_fitter_type, now_fitter_table, now_mes_zm)
        index = ['A'] + index
        # print(index)
        wb = openpyxl.load_workbook(self.lins_file_path + "text_{}.xlsx".format(now_fitter_table))
        ws = wb.active
        # wb_ = openpyxl.load_workbook('banner_complete.xlsx')
        name_fit_list = [x.replace('/', '') for x in now_fitter_list]  # 处理名字特殊符号
        if not os.path.exists(
                self.download_path + '{} {}.xlsx'.format(now_fitter_type, now_fitter_table)):  # 如果文件不存在
            wb_ = openpyxl.Workbook()
            # 处理特殊符号
            wb_.create_sheet('_'.join(name_fit_list))  # 新建一个筛选项的sheet
            ws_ = wb_['_'.join(name_fit_list)]
        else:  # 存在就直接打开
            wb_ = openpyxl.load_workbook(
                self.download_path + '{} {}.xlsx'.format(now_fitter_type, now_fitter_table))
            wb_.create_sheet('_'.join(name_fit_list))  # 新建一个筛选项的sheet
            ws_ = wb_['_'.join(name_fit_list)]
        all_row = ws.max_row  # 获取最大行数
        # print('开始')
        for one_row in range(1, all_row + 1):
            lins_list = []
            for one_index in index:
                # print(one_index, one_row)
                one_value = ws['{}{}'.format(one_index, one_row)].value
                if one_value is None:
                    one_value = ''
                lins_list.append(one_value)
                # print(ws['{}{}'.format(one_index, one_row)].value)
            # print(lins_list)
            ws_.append(lins_list)
        wb_.save(self.download_path + '{} {}.xlsx'.format(now_fitter_type, now_fitter_table))
        # print('写表完成')
        # end = time.time()
        # print('获取并写出无样式的表 时间：%.2f秒' % (end - start))

    def wirte_style(self, now_fitter_type=None, now_fitter_table=None, now_fitter_list=None):
        """
        写样式
        :return:
        """
        # start = time.time()
        zimu = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
        wb = openpyxl.load_workbook(self.download_path + '{} {}.xlsx'.format(now_fitter_type, now_fitter_table))
        name_fit_list = [x.replace('/', '') for x in now_fitter_list]  # 处理名字特殊符号
        ws = wb['_'.join(name_fit_list)]
        z = ws.max_column  # 最大列
        if z <= 26:  # 处理标题的长度对应的字母
            mes_zm = zimu[:z]
            # print(mes_zm)
        else:
            else_zm_len = z - 26
            mes_zm = zimu
            num = 0
            for f_zm in zimu:
                if num >= else_zm_len:
                    break
                for s_zm in zimu:
                    if num >= else_zm_len:
                        break
                    ex_zm = f_zm + s_zm
                    mes_zm.append(ex_zm)
                    num += 1
        Channal_index_list = []  # Channal 的标题
        # else_index_lise = []  # 其他的标题
        if now_fitter_type == 'bybanner':
            all_Channal = ['H/S', 'Hyper', 'Super', 'Mini']
        else:
            all_Channal = ['CBL', 'SCCL', 'ZH']

        title_font = Font(name='Arial', size=10, bold=True)  # 字体
        font = Font(name='Arial', size=10)  # 字体
        align = Alignment(horizontal='center')  # 居中
        fill = PatternFill("solid", fgColor='E5E5E5')  # 标题背景颜色
        shop_fill = PatternFill("solid", fgColor='C1FFC1')  # 标题背景颜色
        color_list = ['27408B', 'FF0000', '4876FF', '63B8FF']  # 其他字体颜色
        # else_title = mes_zm[mes_zm.index(previous_) + 1]  # 获取其他标题
        if now_fitter_type == 'byou' and 'Mini' in now_fitter_table:  # Mini 的只合并这一种
            ws.merge_cells('C1:F1')
        else:
            # 读第二行 判断合并位置
            for one_index in mes_zm[:7]:
                if ws['{}2'.format(one_index)].value in all_Channal:
                    Channal_index_list.append(one_index)
            # 判断其他的标题的开始位置
            previous_ = Channal_index_list[-1]
            else_index = mes_zm[mes_zm.index(previous_) + 1:]
            # else_index = zimu
            else_title = [ws['{}1'.format(x)].value for x in else_index]
            ws.merge_cells('{}1:{}1'.format(Channal_index_list[0], Channal_index_list[-1]))  # 合并Channal
            merge_flag = False  # 合并标志
            for j, one_else_title in enumerate(else_title):
                if one_else_title:
                    if merge_flag == False:
                        lins_list = []
                        lins_list.append(else_index[j])
                        merge_flag = True
                    else:
                        ws.merge_cells('{}1:{}1'.format(lins_list[0], lins_list[-1]))  # 合并其他单元格
                        merge_flag = False
                        lins_list = []
                        lins_list.append(else_index[j])
                        merge_flag = True
                else:
                    lins_list.append(else_index[j])
                if j == len(else_title) - 1:
                    ws.merge_cells('{}1:{}1'.format(lins_list[0], lins_list[-1]))  # 合并其他单元格
        if now_fitter_type == 'bybanner':
            ws['{}1'.format(Channal_index_list[0])] = "Channal"
        else:
            pass
        # 标题格式：居中，加粗,底色
        for one_size in mes_zm:
            if ws['{}1'.format(one_size)].value == None:
                continue
            else:
                ws['{}1'.format(one_size)].font = title_font
                ws['{}1'.format(one_size)].alignment = align
                ws['{}1'.format(one_size)].fill = fill
        # 门店数量格式：字体居中，底色
        for one_size in mes_zm:
            ws['{}3'.format(one_size)].font = font
            ws['{}3'.format(one_size)].alignment = align
            ws['{}3'.format(one_size)].fill = shop_fill
        # 其他位置格式：字体居中
        for i in range(2, ws.max_row + 1):
            if '满分' in str(ws['A{}'.format(i)].value):  # 特殊行数要颜色
                if i == 5:
                    font = Font(name='Arial', size=10, color=color_list[0], bold=True)
                elif i == 6:
                    font = Font(name='Arial', size=10, color=color_list[1], bold=True)
                elif i == 7:
                    font = Font(name='Arial', size=10, color=color_list[2], bold=True)
                elif i == 8:
                    font = Font(name='Arial', size=10, color=color_list[3], bold=True)
                else:
                    font = Font(name='Arial', size=10, color=color_list[2], bold=True)
            else:
                font = Font(name='Arial', size=10)  # 普通字体
            for one_size in mes_zm:
                ws['{}{}'.format(one_size, i)].font = font
                ws['{}{}'.format(one_size, i)].alignment = align
        wb.save(self.download_path + '{} {}.xlsx'.format(now_fitter_type, now_fitter_table))
        # print('样式完成')
        # end = time.time()
        # print('写样式 时间：%.2f秒' % (end - start))

    def all_down(self, now_fitter_type, now_fitter_table, values):
        for one_fit, fit_value in values.items():
            if now_fitter_type == 'bybanner':  # banner 有3项
                for two_fit, two_fit_value in fit_value.items():
                    for three_fit in two_fit_value:
                        try:
                            self.complete_num += 1
                            self.count_out.emit(self.complete_num, 'bybanner')
                            fitter_list = [one_fit, two_fit, three_fit]
                            print(now_fitter_table)
                            print(fitter_list)
                            start = time.time()
                            mes_zm = self.get_fitter_table(now_fitter_type, now_fitter_table, fitter_list)  # 生成基础表
                            self.get_complete_table(now_fitter_type, now_fitter_table, fitter_list, mes_zm)  # 生成无样式的完整表
                            self.wirte_style(now_fitter_type, now_fitter_table, fitter_list)  # 写样式
                            end = time.time()
                            print('生成总时间：%.2f秒' % (end - start))
                        except Exception as e:
                            print('错误')
                            print(e)
                            nt = datetime.datetime.now().strftime('%Y.%m.%d %H.%M.%S')
                            with open('log.txt', 'a') as f:
                                f.write('{} ======== {}'.format(str(nt), str(e)))
            else:  # ou 只有2项
                for two_fit in fit_value:
                    try:
                        self.complete_num += 1
                        self.count_out.emit(self.complete_num, 'byou')
                        fitter_list = [one_fit, two_fit]
                        print(now_fitter_table)
                        print(fitter_list)
                        start = time.time()
                        mes_zm = self.get_fitter_table(now_fitter_type, now_fitter_table, fitter_list)  # 生成基础表
                        self.get_complete_table(now_fitter_type, now_fitter_table, fitter_list, mes_zm)  # 生成无样式的完整表
                        self.wirte_style(now_fitter_type, now_fitter_table, fitter_list)  # 写样式
                        end = time.time()
                        print('生成总时间：%.2f秒' % (end - start))
                    except Exception as e:
                        print(e)
                        input('错误')

    def run(self):
        '''
        生成完整的表
        :return:
        '''
        # banner
        self.all_start = time.time()
        if self.fitter_type == 'bybanner':
            start = time.time()
            pool = ThreadPoolExecutor(max_workers=10)  # 设置线程数
            data_list = []
            for key, values in self.banner_table_dict.items():
                data_list.append([key, values])
            while data_list:
                try:
                    one_data = data_list.pop(0)
                    key = one_data[0]
                    values = one_data[1]
                    self.fitter_table = key.replace('.src', '')  # 获取表名
                    # print(self.fitter_table)
                    self.all_down(self.fitter_type, self.fitter_table, values)
                except Exception as e:
                    print('错误')
                    print(e)
                    nt = datetime.datetime.now().strftime('%Y.%m.%d %H.%M.%S')
                    with open('log.txt', 'a') as f:
                        f.write('{} ========大错误 {}'.format(str(nt), str(e)))
                # pool.submit(self.all_down, self.fitter_type, self.fitter_table, values)
                # self.get_fitter_table()  # 生成基础表
                # self.get_complete_table()  # 生成无样式的完整表
                # self.wirte_style()  # 写样式
            # pool.shutdown()

        else:
            start = time.time()
            pool = ThreadPoolExecutor(max_workers=10)  # 设置线程数
            data_list = []
            for key, values in self.ou_table_dict.items():
                data_list.append([key, values])
            while data_list:
                one_data = data_list.pop(0)
                key = one_data[0]
                values = one_data[1]
                self.fitter_table = key.replace('.src', '')  # 获取表名
                print(self.fitter_table)
                self.all_down(self.fitter_type, self.fitter_table, values)
                # pool.submit(self.all_down, self.fitter_type, self.fitter_table, values)
            # pool.shutdown()
        # #ou
        # for key, values in self.ou_table_dict.items():
        #     self.fitter_type = 'byou'
        #     self.fitter_table = key.replace('.src', '')  # 获取表名
        #     print(self.fitter_table)
        #     for one_fit, fit_value in values.items():
        #         for two_fit in fit_value:
        #             self.fitter_list = [one_fit, two_fit]
        #             self.get_fitter_table()  # 生成基础表
        #             self.get_complete_table()  # 生成无样式的完整表
        #             self.wirte_style()  # 写样式
        try:
            print(self.all_end - self.all_start)
            self.all_end = time.time()
            self.str_out.emit('生成总时间：%.2f秒' % (self.all_end - self.all_start), self.fitter_type)
        except:
            pass
        print('完成')
        # self.str_out.emit('下载完成：{} {}.xlsx'.format(self.fitter_type, self.fitter_table), self.fitter_type)
        # end = time.time()
        # print(end - start)


class VisitMain(object):
    """界面"""

    def __init__(self):
        self.summary_file = '测试设置表.xlsx'
        self.bannerS_file = 'bybannerSS.xlsx'
        self.ouS_file = 'byouSS.xlsx'
        self.app = QtWidgets.QApplication(sys.argv)
        self.wait_window = QtWidgets.QWidget()
        self.wait_ui = Wait_Form()
        self.wait_window.setFont(QFont("Microsoft YaHei", 9))
        self.wait_ui.setupUi(self.wait_window)
        self.wait_window.setWindowFlags(Qt.CustomizeWindowHint)
        self.wait_window.show()
        self.get_all_data()
        self.all_table = ['Total score.Hyper', 'Total score.Super', 'Total score.Mini',
                          'Availability.HS', 'Availability.Mini', 'SOVI', 'Cooler', '2nd Display',
                          'Thematic Display', 'Price']
        self.transpose_table_list = ['SKU Availability by Store', 'SKU Facing by Store', 'SKU Price by Store',
                                     'SOVI by Store', 'Cooler by Store', '2nd Display by Store', 'Thematic by Store']
        self.summary_file = '测试设置表.xlsx'
        self.table_banner_dict = {}  # banner 筛选框选项
        self.table_ou_dict = {}  # ou 筛选框选项
        self.main_window = QtWidgets.QWidget()
        self.ui = Ui_Form()
        self.ui.setupUi(self.main_window)
        self.main_window.setWindowIcon(QIcon('./system_File/图标.png'))
        self.main_window.setFont(QFont("Microsoft YaHei", 9))
        self.get_fitter_list()  # 初始化筛选
        print('初始化数据中')
        # self.wait = WaitWindow()
        # self.get_file()  # 初始化所有表
        self.connect()
        sys.exit(self.app.exec_())

    def connect(self):
        self.ui.pushButton.clicked.connect(self.get_fitter_list)
        self.ui.comboBox.currentIndexChanged.connect(self.banner_select_st_change)  # 表 筛选框
        self.ui.comboBox_2.currentIndexChanged.connect(self.banner_select_nd_change)  # BG 筛选框
        self.ui.comboBox_3.currentIndexChanged.connect(self.banner_select_rd_change)  # OU 筛选框
        self.ui.comboBox_9.currentIndexChanged.connect(self.ou_select_st_change)  # channel
        self.ui.comboBox_10.currentIndexChanged.connect(self.ou_select_nd_change)  # banner
        self.ui.pushButton.clicked.connect(self.start_download_banner)  # banner 下载
        self.ui.pushButton_2.clicked.connect(self.start_show_banner)  # banner 查看
        self.ui.pushButton_5.clicked.connect(self.start_download_ou)  # ou 下载
        self.ui.pushButton_6.clicked.connect(self.start_show_ou)  # ou查看
        self.ui.pushButton_3.clicked.connect(self.down_allbanner_table)  # 下载所有banner表
        self.ui.pushButton_4.clicked.connect(self.down_allou_table)  # 下载所有ou表

    # 配置筛选框
    def banner_select_st_change(self):
        """更改第一层选择框(bybanner)"""
        tag_name = self.ui.comboBox.currentText() + '.src'
        self.ui.comboBox_2.clear()
        self.ui.comboBox_3.clear()
        self.ui.comboBox_4.clear()
        else_selector = self.table_banner_dict.get(tag_name)
        if else_selector:
            for one in else_selector:  # BG
                self.ui.comboBox_2.addItem(one)

    def banner_select_nd_change(self):
        try:
            tag_name = self.ui.comboBox.currentText() + '.src'
            second_name = self.ui.comboBox_2.currentText()
            self.ui.comboBox_3.clear()
            self.ui.comboBox_4.clear()
            else_selector = self.table_banner_dict.get(tag_name).get(second_name)
            if else_selector:
                # print(else_selector)
                for one in else_selector:  # BG
                    self.ui.comboBox_3.addItem(one)
        except:
            pass

    def banner_select_rd_change(self):
        try:
            tag_name = self.ui.comboBox.currentText() + '.src'
            second_name = self.ui.comboBox_2.currentText()
            third_name = self.ui.comboBox_3.currentText()
            self.ui.comboBox_4.clear()
            if self.table_banner_dict.get(tag_name).get(second_name) is None:
                return
            else_selector = self.table_banner_dict.get(tag_name).get(second_name).get(third_name)
            if else_selector:
                for one in else_selector:  # BG
                    self.ui.comboBox_4.addItem(one)
        except:
            pass

    def ou_select_st_change(self):
        tag_name = self.ui.comboBox_9.currentText() + '.src'
        self.ui.comboBox_10.clear()
        self.ui.comboBox_11.clear()
        else_selector = self.table_ou_dict.get(tag_name)
        if else_selector:
            for one in else_selector:
                self.ui.comboBox_10.addItem(one)

    def ou_select_nd_change(self):
        try:
            tag_name = self.ui.comboBox_9.currentText() + '.src'
            second_name = self.ui.comboBox_10.currentText()
            self.ui.comboBox_11.clear()
            else_selector = self.table_ou_dict.get(tag_name).get(second_name)
            if else_selector:
                for one in else_selector:
                    self.ui.comboBox_11.addItem(one)
        except:
            pass

    # 初始化筛选项
    def get_fitter_list(self):
        """获取筛选框内容"""
        df = pd.read_excel(self.summary_file, sheet_name='TableCopy')  # 可以通过sheet_name来指定读取的表单
        all_title_name = df.columns.values  # 获取所有表头
        table_nameid_title = all_title_name[10:]
        table_nameid = df[table_nameid_title]
        table_nameid_list = list()  # ou 表id
        table_nameid_banner_list = list()  # banner 表id
        for j, one_src in enumerate(table_nameid):
            lins_list = list(table_nameid[one_src].dropna().values)
            lins_list.insert(0, table_nameid_title[j])
            table_nameid_list.append(lins_list)
            table_nameid_banner_list.append([table_nameid_title[j]])
        # print(table_nameid_list)
        ##获取banner 的表id##
        byBanner_df = pd.read_excel(self.summary_file, sheet_name='bybanner')  # 获取表byBanner 的数据
        all_title_name = byBanner_df.columns.values  # 获取所有表头
        table_row = byBanner_df[all_title_name[:5]]
        max_row = list(table_row.iloc[len(table_row) - 1])[-1]
        for one_table in table_nameid_banner_list:
            lins_list = [x for x in range(1, max_row + 1)]
            for i in lins_list:
                one_table.append(i)
        # print(table_nameid_banner_list)
        # bybanner表的筛选框
        banner_df = pd.read_excel(self.summary_file, sheet_name='fiterbybanner')
        for one_table in table_nameid_banner_list:  # 其中一个id
            lins_data_list = []
            for one_id in one_table[1:]:  # 分开筛选框#
                fitter_df = banner_df[banner_df.FilterNo.isin([one_id])]
                for data in fitter_df.iterrows():  # 其中一个id的
                    lins_data_list.append([data[1]['BG'], data[1]['OU'], data[1]['City']])
            data_dict = {}
            # 生成每层对应的选项
            for item in lins_data_list:
                t1 = item[0]
                t2 = item[1]
                t3 = item[2]
                if t1 in data_dict:
                    t1_dict = data_dict[t1]
                    if t2 in t1_dict:
                        data_dict[t1][t2].append(t3)
                    else:
                        t2_dict = [t3]
                        data_dict[t1].update({
                            t2: t2_dict
                        })
                else:
                    t2_dict = [t3]
                    data_dict[t1] = {
                        t2: t2_dict
                    }

                # for one_row in fitter_df.iterrows():
                #     one_.append(one_row[1]["BG"])
                #     two_.append(one_row[1]["OU"])
                #     three_.append(one_row[1]["City"])
                # self.table_banner_dict[one_table[0]] = [list(set(one_)), list(set(two_)), list(set(three_))]
            try:
                data_dict['Total']['Total'] = ['Total']
            except Exception as e:
                print(e)
                pass
            self.table_banner_dict[one_table[0]] = data_dict
        # # byou表的筛选框
        ou_df = pd.read_excel(self.summary_file, sheet_name='fiterbyou')
        for one_table in table_nameid_list:
            lins_data_list = []
            for one_id in one_table[1:]:  # 分开筛选框#
                fitter_df = ou_df[ou_df.FilterNo.isin([one_id])]
                for one_row in fitter_df.iterrows():
                    lins_data_list.append([one_row[1]['Channel'], one_row[1]['Banner']])
            data_dict = {}
            # 生成每层对应的选项
            for item in lins_data_list:
                t1 = item[0]
                t2 = item[1]
                if t1 in data_dict:
                    t1_list = data_dict[t1]
                    if t2 in t1_list:
                        pass
                    else:
                        t1_list.append(t2)
                    data_dict[t1] = t1_list
                else:
                    data_dict[t1] = [t2]
            self.table_ou_dict[one_table[0]] = data_dict
        # banner
        # self.ui.comboBox.addItem('选择表')
        for word in self.all_table:
            self.ui.comboBox.addItem(word)
        for word in self.transpose_table_list:
            self.ui.comboBox.addItem(word)
        # ou
        # self.ui.comboBox_9.addItem('选择表')
        for word in self.all_table:
            self.ui.comboBox_9.addItem(word)
        for word in self.transpose_table_list:
            self.ui.comboBox_9.addItem(word)

    # banner操作
    def start_download_banner(self):
        """banner 下载"""
        table_type = 'bybanner'
        table_name = self.ui.comboBox.currentText()  # 表名
        second_name = self.ui.comboBox_2.currentText()
        third_name = self.ui.comboBox_3.currentText()
        four_name = self.ui.comboBox_4.currentText()
        self.banner_def = GenerateTable(table_type, table_name, [second_name, third_name, four_name])
        self.banner_def.start()
        # self.banner_def.download_excel()
        self.banner_def.str_out.connect(self.pri_text)

    def start_show_banner(self):
        """banner 展示"""
        table_type = 'bybanner'
        table_name = self.ui.comboBox.currentText()  # 表名
        if 'by Store' in table_name:  # by store
            second_name = ''
            third_name = ''
            four_name = ''
        else:
            second_name = self.ui.comboBox_2.currentText()
            third_name = self.ui.comboBox_3.currentText()
            four_name = self.ui.comboBox_4.currentText()
        self.banner_def = GenerateTableShow(table_type, table_name, [second_name, third_name, four_name])
        self.banner_def.start()
        # self.banner_def.download_excel()
        self.banner_def.str_out.connect(self.pri_text)
        self.banner_def.title_out.connect(self.update_title)  # 刷标题
        self.banner_def.data_out.connect(self.update_data)  # 刷数据
        self.banner_def.clear_out.connect(self.clear_win)  # 清除数据
        self.banner_def.clear_str.connect(self.clear_str)  # 清除打印窗口

    # ou操作
    def start_download_ou(self):
        """banner 下载"""
        table_type = 'byou'
        table_name = self.ui.comboBox_9.currentText()  # 表名
        second_name = self.ui.comboBox_10.currentText()
        third_name = self.ui.comboBox_11.currentText()
        self.banner_def = GenerateTable(table_type, table_name, [second_name, third_name])
        self.banner_def.start()
        # self.banner_def.download_excel()
        self.banner_def.str_out.connect(self.pri_text)

    def start_show_ou(self):
        """banner 展示"""
        table_type = 'byou'
        table_name = self.ui.comboBox_9.currentText()  # 表名
        if 'by Store' in table_name:  # by store
            second_name = ''
            third_name = ''
        else:
            second_name = self.ui.comboBox_10.currentText()
            third_name = self.ui.comboBox_11.currentText()
        self.banner_def = GenerateTableShow(table_type, table_name, [second_name, third_name])
        self.banner_def.start()
        # self.banner_def.download_excel()
        self.banner_def.str_out.connect(self.pri_text)
        self.banner_def.title_out.connect(self.update_title)  # 刷标题
        self.banner_def.data_out.connect(self.update_data)  # 刷数据
        self.banner_def.clear_out.connect(self.clear_win)  # 清除数据
        self.banner_def.clear_str.connect(self.clear_str) #清除打印窗口

    # 更新数据
    def update_data(self, list, t_type):
        """刷数据"""
        if t_type == 'bybanner':
            row_count = self.ui.tableWidget.rowCount()
            self.ui.tableWidget.insertRow(row_count)
            for j, one_data in enumerate(list):
                self.ui.tableWidget.setItem(row_count, j, QTableWidgetItem(str(one_data)))
                self.ui.tableWidget.scrollToTop()  # 保持滚动条最上
            for i in range(len(list)):  # 刷居中
                self.ui.tableWidget.item(row_count, i).setTextAlignment(Qt.AlignCenter)
        else:
            row_count = self.ui.tableWidget_3.rowCount()
            self.ui.tableWidget_3.insertRow(row_count)
            for j, one_data in enumerate(list):
                self.ui.tableWidget_3.setItem(row_count, j, QTableWidgetItem(str(one_data)))
                self.ui.tableWidget_3.scrollToTop()  # 保持滚动条最上
            for i in range(len(list)):  # 刷居中
                self.ui.tableWidget_3.item(row_count, i).setTextAlignment(Qt.AlignCenter)

    def update_title(self, title_list, t_type):
        if t_type == "bybanner":
            self.ui.tableWidget.setColumnCount(len(title_list))
            try:
                self.ui.tableWidget.setHorizontalHeaderLabels(title_list)
            except Exception as e:
                print(e)
            # print('成功')
        else:
            self.ui.tableWidget_3.setColumnCount(len(title_list))
            self.ui.tableWidget_3.setHorizontalHeaderLabels(title_list)

    # 读取数据库获取所有值
    def get_all_data(self):
        self.start_ = GetAllTable()
        self.start_.start()
        self.start_.stop_flag.connect(self.closewin)

    def closewin(self):
        self.wait_window.close()
        self.main_window.show()

    def clear_win(self, t_type):
        """清屏"""
        if t_type == 'bybanner':
            row_count = self.ui.tableWidget.rowCount()
            for i in range(row_count):  # 清空界面
                self.ui.tableWidget.removeRow(0)
        else:
            row_count = self.ui.tableWidget_3.rowCount()
            for i in range(row_count):  # 清空界面
                self.ui.tableWidget_3.removeRow(0)

    def pri_text(self, text, t_type):
        """填写窗口信息"""
        if t_type == 'bybanner':
            self.ui.textEdit.append(text)
        else:
            self.ui.textEdit_2.append(text)

    def clear_str(self,t_type):
        if t_type == 'bybanner':
            self.ui.textEdit.clear()
        else:
            self.ui.textEdit_2.clear()

    # 导出所有banner表
    def down_allbanner_table(self):
        '''下载所有数据'''
        print(self.table_ou_dict)
        print(self.table_banner_dict)
        self.allbanner_file_d = DownloadAllFiles('bybanner', '', [], self.table_banner_dict,
                                                 self.table_ou_dict)
        try:
            self.allbanner_file_d.start()
            # self.banner_def.download_excel()
            self.allbanner_file_d.str_out.connect(self.pri_text)
            self.allbanner_file_d.count_out.connect(self.count_num)
        except Exception as e:
            print(e)

    def down_allou_table(self):
        '''下载所有数据'''
        # print(self.table_ou_dict)
        # print(self.table_banner_dict)
        self.allou_file_d = DownloadAllFiles('byou', '', [], self.table_banner_dict,
                                             self.table_ou_dict)
        try:
            self.allou_file_d.start()
            # self.banner_def.download_excel()
            self.allou_file_d.str_out.connect(self.pri_text)
            self.allou_file_d.count_out.connect(self.count_num)
        except Exception as e:
            print(e)

    def count_num(self, num, t_type):
        if t_type == 'bybanner':
            self.ui.textEdit.clear()
            self.ui.textEdit.append("生成表数：{}（共约1760个）".format(num))
        else:
            self.ui.textEdit_2.clear()
            self.ui.textEdit_2.append("生成表数：{}（共约400个）".format(num))


class WaitWindow(object):
    """等待界面"""

    def __init__(self):
        self.bannerS_file = 'bybannerSS.xlsx'
        self.ouS_file = 'byouSS.xlsx'
        # self.app = QtWidgets.QApplication(sys.argv)
        self.main_window = QtWidgets.QMainWindow()
        self.ui = Wait_Form()
        self.app = QtWidgets.QApplication(sys.argv)
        self.main_window.setFont(QFont("Microsoft YaHei", 9))
        self.ui.setupUi(self.main_window)
        self.main_window.setWindowFlags(Qt.CustomizeWindowHint)
        self.main_window.show()
        self.get_all_data()
        sys.exit(self.app.exec_())

    def get_all_data(self):
        self.start_ = GetAllTable()
        self.start_.start()
        self.start_.stop_flag.connect(self.closewin)

    def closewin(self):
        self.main_window.close()


class GetAllTable(QThread):
    """读取数据库的数据"""
    stop_flag = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.bannerS_file = 'bybannerSS.xlsx'
        self.ouS_file = 'byouSS.xlsx'

    def run(self):
        global all_banner_data
        global all_ou_data
        start = time.time()
        for one_table in all_table:
            self.all_data(one_table)
        end = time.time()
        print(end - start)
        self.stop_flag.emit()

    def all_data(self, one_fitter):
        self.db = sqlite3.connect('./data.db')
        self.cur = self.db.cursor()
        print('banner:' + one_fitter)
        read_sql = """SELECT * FROM `{}`""".format('b' + one_fitter + '.src')
        self.cur.execute(read_sql)  # 执行sql语句
        row = self.cur.fetchall()
        table_fitter_pd = pd.DataFrame(list(row))
        last_title = list(table_fitter_pd.columns.values)[-1]
        table_fitter_pd.rename(columns={last_title: 'tid'}, inplace=True)
        for one_title in list(table_fitter_pd.columns.values)[1:]:
            table_fitter_pd[one_title] = table_fitter_pd[one_title].astype('float')
        all_banner_data[one_fitter] = table_fitter_pd
        print('ou: ' + one_fitter)
        read_sql = """SELECT * FROM `{}`""".format('o' + one_fitter + '.src')
        self.cur.execute(read_sql)  # 执行sql语句
        row = self.cur.fetchall()
        table_fitter_pd = pd.DataFrame(list(row))
        last_title = list(table_fitter_pd.columns.values)[-1]
        table_fitter_pd.rename(columns={last_title: 'tid'}, inplace=True)
        for one_title in list(table_fitter_pd.columns.values)[1:]:
            table_fitter_pd[one_title] = table_fitter_pd[one_title].astype('float')
        all_ou_data[one_fitter] = table_fitter_pd


if __name__ == '__main__':
    # a = GenerateTable()
    # # a.get_banner_fitter_id()
    # a.get_fitter_table()
    # a.get_complete_table()
    # a.wirte_style()
    VisitMain()
    # a = GetAllTable()
    # a.text()
