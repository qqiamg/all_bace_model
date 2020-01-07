import sqlite3
import openpyxl


###保存汇总表到sqlite###

class SaveInSql(object):
    """保存进数据库"""

    def __init__(self):
        self.db = sqlite3.connect('./data.db')  # 链接数据库，不存在就创建
        self.cur = self.db.cursor()
        self.all_title = []

    def save_sql(self, type):
        """
        保存banner和ou到数据库
        :return:
        """
        if type == 'bybanner':
            wb = openpyxl.load_workbook('bybannerSS.xlsx')
            first_zm = 'b'
        else:
            wb = openpyxl.load_workbook('byouSS.xlsx')
            first_zm = 'o'
        all_sheet_name = wb.sheetnames  # 所有sheet名字
        print(type)
        for sheet_name in all_sheet_name:
            ws = wb[sheet_name]
            max_len = ws.max_column  # 最大列数
            # 获取标题
            rows = ws.rows
            data_list = []
            print(sheet_name)
            self.all_title = []
            for j, row in enumerate(rows):
                if j == 0:  # 第一行，读表头，建表
                    lins_list = []
                    for i in range(max_len):
                        one_title = "'" + str(row[i].value) + "'" + ' TEXT(255)'
                        lins_list.append(one_title)
                        self.all_title.append("'" + str(row[i].value) + "'")
                    all_title = ','.join(lins_list)
                    del_sql = """DROP TABLE IF EXISTS `{}`""".format(first_zm + sheet_name)  # 删除
                    self.cur.execute(del_sql)  # 执行sql语句(删除)
                    title_sql = """CREATE TABLE `{}` ({})""".format(first_zm + sheet_name, all_title)  # 创建
                    # print(title_sql)
                    self.cur.execute(title_sql)  # 执行sql语句(建表)
                    self.db.commit()
                else:  # 整理数据
                    lins_list = []
                    for i in range(max_len):
                        if row[i].value or row[i].value == 0:
                            lins_list.append('"' + str(row[i].value) + '"')
                        else:
                            lins_list.append('null')
                    one_line_data = '(' + ','.join(lins_list) + ')'
                    data_list.append(one_line_data)
            all_data = ','.join(data_list)
            data_sql = """INSERT INTO `{}` ({}) VALUES {}""".format(first_zm + sheet_name, ','.join(self.all_title),
                                                                    all_data)
            # print(data_sql)
            self.cur.execute(data_sql)  # 执行sql语句(插数据)
            self.db.commit()
            # break

    def run(self):
        self.save_sql('bybanner')
        self.save_sql('byou')


if __name__ == '__main__':
    a = SaveInSql()
    a.run()
