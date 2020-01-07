import os


import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

from main_windows import Ui_Form


# from statement import Statement

###获取筛选项并整理到一个excel###

class GetFilter(object):

    def __init__(self):
        self.path_list = []
        self.bace_path = r'D:\杨伟强\工作\可口可乐数据处理\table' #获取所有路径下的文件

    def get_bybanner_city(self):
        """获取bybanner city(测试去重excel.xlsx)"""
        lins_df = pd.read_excel('pop.xlsx')
        lins_df['tag'] = lins_df['BG'] + lins_df['OU'] + lins_df['City']
        new_lins_df = lins_df[['BG', 'OU', 'City', 'tag']]
        print(new_lins_df)
        new_lins_df.drop_duplicates(subset='tag', keep='first', inplace=True)
        new_lins_df.sort_values(by=["BG", "OU", "City"], inplace=True)
        b = new_lins_df.iloc[:, :3]
        b.index = b.index + 1  # shifting index
        df = pd.DataFrame(data=[["Total", "Total", "Total"]], columns=["BG", "OU", "City"])
        com_db = df.append(b)
        # b.loc[-1] = ["Total", "Total", "Total"]
        # b = b.sort_index()
        print(com_db)
        com_db.to_excel('测试去重excel.xlsx', sheet_name='测试', index=False)

    def get_allfile_path(self, sPath):
        """列出当前路径下的所有文件夹和文件　并进行遍历"""
        for schild in os.listdir(sPath):
            # 拼接地址
            sChildPath = os.path.join(sPath, schild)
            # 判断当前遍历到的是文件还是文件夹
            if os.path.isdir(sChildPath):
                # 再次递归调用
                self.get_allfile_path(sChildPath)
            else:
                if '%' in sChildPath or '~$' in sChildPath:  # 除去不需要的
                    pass
                else:
                    sChildPath_list = sChildPath.rsplit("\\", 1)
                    sChildPath_list.append(sChildPath)
                    self.path_list.append(sChildPath_list)

    def save(self):
        """获取所有路径下的文件(测试路径总和.xlsx)"""
        bace_path = self.bace_path
        self.get_allfile_path(bace_path)
        # writer = pd.ExcelWriter("测试路径总和.xlsx")
        df = pd.DataFrame(data=self.path_list, columns=["PathName", "FileName", "all_path"])
        df.sort_values(by=["PathName", "FileName"], inplace=True)
        print(df)
        # df.to_excel(writer, index=False, sheet_name="Dirs")
        df.to_excel('测试路径总和.xlsx', sheet_name='测试', index=False)
        # writer.save()
        # writer.close()

    def get_bybanner_ou_No(self):
        """获取bybanner_No 和 ou_No(测试banner_ou总和.xlsx)"""
        lins_df = pd.read_excel('测试路径总和.xlsx')
        banner_df = pd.read_excel('fiter_bybanner.xlsx', sheet_name='fiterbybanner')
        ou_df = pd.read_excel('fiter_byou.xlsx', sheet_name='fiterbyou')
        # print(lins_df)
        writer = pd.ExcelWriter("测试banner_ou总和.xlsx")
        banner_list = []
        ou_list = []
        # banner_no = 1
        # ou_no = 1
        # 处理 fiter_bybanner 数据
        banner_fit_name = list(banner_df['Combination'])
        banner_fit_No = list(banner_df['FilterNo'])
        # 处理 fiter_ou 数据
        ou_fit_name = list(ou_df['Combination'])
        ou_fit_No = list(ou_df['FilterNo'])
        # print(banner_fit_name)
        print(ou_fit_name)
        for one in lins_df.iterrows():
            if "by banner" in one[1]['PathName']:  # 为bybanner的路径
                type = one[1]['FileName'].replace('可乐-Banner-', '').replace('(Abs).xlsx', '')
                lins_list = []
                lins_list.append(one[1]["all_path"])  # 文件路径
                if 'base=BG' in one[1]['PathName']:  # BG 分类的
                    lins_list.append(type)
                    lins_list.append('Total')
                    lins_list.append('Total')
                elif 'base=City' in one[1]['PathName']:  # City 分类的
                    lins_list.append('Total')
                    lins_list.append('Total')
                    lins_list.append(type)
                else:  # OU 分类的
                    lins_list.append('Total')
                    lins_list.append(type)
                    lins_list.append('Total')
                # print(''.join(lins_list[1:]))
                if ''.join(lins_list[1:]) in banner_fit_name:
                    index_ = banner_fit_name.index(''.join(lins_list[1:]))
                    lins_list.append(banner_fit_No[index_])
                else:
                    continue
                # lins_list.append(banner_no)
                bacontinue_flag = False
                for i in banner_list:  # 已存在就不加入了
                    if i[-1] == lins_list[-1]:
                        bacontinue_flag = True
                        break
                if bacontinue_flag:
                    continue
                banner_list.append(lins_list)
                # banner_no += 1
            elif "by bottler" in one[1]['PathName']:
                lins_list = []
                lins_list.append(one[1]["all_path"])  # 文件路径
                if 'base=Channel' in one[1]['PathName']:  # BG 分类的
                    type = one[1]['FileName'].replace('可乐-BG&OU-', '').replace('(Abs).xlsx', '')
                    if type == 'H S':
                        type = 'H/S'
                    lins_list.append(type)
                    lins_list.append('Total')

                elif 'base=Banner' in one[1]['PathName']:  # City 分类的
                    type = one[1]['FileName'].replace('可乐-BG&OU-', '').replace('(Abs).xlsx', '').split('_')
                    if type[0] == 'H S':
                        type[0] = 'H/S'
                    lins_list.append(type[0])
                    lins_list.append(type[1])
                print(''.join(lins_list[1:]))
                if ''.join(lins_list[1:]) in ou_fit_name:
                    index_ = ou_fit_name.index(''.join(lins_list[1:]))
                    lins_list.append(ou_fit_No[index_])
                else:
                    continue
                # lins_list.append(ou_no)
                continue_flag = False
                for i in ou_list:  # 已存在就不加入了
                    if i[-1] == lins_list[-1]:
                        continue_flag = True
                        break
                if continue_flag:
                    continue
                ou_list.append(lins_list)
                # ou_no += 1

        df_banner = pd.DataFrame(data=banner_list, columns=["Source File", "BG", "OU", "City", "NO."])
        df_banner.sort_values(by=["NO."], inplace=True)
        df_banner.to_excel(writer, index=False, sheet_name="bybanner")
        df_ou = pd.DataFrame(data=ou_list, columns=["Source File", "Channel", "Banner", "NO."])
        df_ou.sort_values(by=["NO."], inplace=True)
        df_ou.to_excel(writer, index=False, sheet_name="byou")
        writer.save()
        writer.close()
        # df.sort_values(by=["Source File"], inplace=True)
        # print(df_banner)

    def get_bybanner_fiter_no(self):
        """获取bybanner_fiter No.(测试fiter_bybanner.xlsx)"""
        lins_df = pd.read_excel('测试去重excel.xlsx')
        all_list = []
        for one_df in lins_df.iterrows():
            now_list = [one_df[1]["BG"], one_df[1]["OU"], one_df[1]["City"]]
            # print(now_list)
            all_list.append(now_list)
        j = 2
        ouNo = 1
        a_ouNo = 1
        maxFltNo = 1
        all_fit_list = []
        print(len(all_list))
        for i in range(len(all_list)):
            lins_list = ['Total', 'Total', all_list[i][2]]
            com_name = 'Total' + 'Total' + all_list[i][2]
            lins_list.append(com_name)
            lins_list.append(maxFltNo)
            city_flt = maxFltNo
            all_fit_list.append(lins_list)
            j += 1

            if city_flt > 1:
                lins_list = ['Total', all_list[i][1]]
                print(i)
                if all_list[i - 1][1] != all_list[i][1]:
                    lins_list.append('Total')
                    com_name = 'Total' + all_list[i][1] + 'Total'
                    lins_list.append(com_name)
                    try:
                        if all_list[i + 1][1] == all_list[i][1]:
                            maxFltNo += 1
                    except:
                        pass
                    lins_list.append(maxFltNo)
                    all_fit_list.append(lins_list)
                    a_ouNo = maxFltNo
                    ouNo += 1
                    j += 1
                    lins_list = ['Total', all_list[i][1], all_list[i][2]]
                    com_name = 'Total' + all_list[i][1] + all_list[i][2]
                    lins_list.append(com_name)
                    lins_list.append(city_flt)
                    all_fit_list.append(lins_list)
                else:
                    lins_list.append(all_list[i][2])
                    com_name = 'Total' + all_list[i][1] + all_list[i][2]
                    lins_list.append(com_name)
                    lins_list.append(city_flt)
                    all_fit_list.append(lins_list)
                j += 1
                if all_list[i - 1][0] != all_list[i][0]:
                    lins_list = [all_list[i][0], 'Total', 'Total']
                    com_name = all_list[i][0] + 'Total' + 'Total'
                    lins_list.append(com_name)
                    try:
                        if all_list[i + 1][0] == all_list[i][0]:
                            maxFltNo += 1
                    except:
                        pass
                    lins_list.append(maxFltNo)
                    all_fit_list.append(lins_list)
                    j += 1
                if all_list[i - 1][1] != all_list[i][1]:
                    lins_list = [all_list[i][0], all_list[i][1], 'Total']
                    com_name = all_list[i][0] + all_list[i][1] + 'Total'
                    lins_list.append(com_name)
                    try:
                        if all_list[i + 1][1] == all_list[i][1]:
                            lins_list.append(a_ouNo)
                        else:
                            lins_list.append(city_flt)
                    except:
                        lins_list.append(city_flt)
                    all_fit_list.append(lins_list)
                    j += 1
                    lins_list = [all_list[i][0], all_list[i][1], all_list[i][2]]
                    com_name = all_list[i][0] + all_list[i][1] + all_list[i][2]
                    lins_list.append(com_name)
                    lins_list.append(city_flt)
                    all_fit_list.append(lins_list)
                else:
                    lins_list = [all_list[i][0], all_list[i][1], all_list[i][2]]
                    com_name = all_list[i][0] + all_list[i][1] + all_list[i][2]
                    lins_list.append(com_name)
                    lins_list.append(city_flt)
                    all_fit_list.append(lins_list)
            maxFltNo += 1
        print(all_fit_list)
        df_ou = pd.DataFrame(data=all_fit_list, columns=['BG', 'OU', 'City', 'Combination', 'FilterNo'])
        df_ou.to_excel('fiter_bybanner.xlsx', index=False, sheet_name='fiterbybanner')

    def get_byou_fiter_no(self):
        """待整理逻辑"""
        # 获取各类超市的规模
        all_market_df = pd.read_excel('pop.xlsx')
        market_type = ['Hyper', 'Super', 'Mini']
        lins_df = all_market_df[['Channel', 'Banner']]
        select_groupby = lins_df.groupby(['Banner', 'Channel'])
        all_name = []
        all_dict = dict()  # 归类统计
        for name, data in select_groupby:
            lins_list = [0, 0, 0]
            if name[1] != 'Hyper' and name[1] != 'Super' and name[1] != 'MINI':  # 剔除不要的
                continue
            if name[0] not in all_name:
                all_name.append(name[0])
                if name[1] == 'Hyper':
                    lins_list[0] = len(data)
                elif name[1] == 'Super':
                    lins_list[1] = len(data)
                else:
                    lins_list[2] = len(data)
                lins_list.insert(0, lins_list[0] + lins_list[1])
                lins_list.insert(0, lins_list[1] + lins_list[2] + lins_list[3])
                all_dict.update({name[0]: lins_list})
            else:
                ard_list = all_dict[name[0]]
                if name[1] == 'Hyper':
                    ard_list[2] = len(data)
                elif name[1] == 'Super':
                    ard_list[3] = len(data)
                else:
                    ard_list[4] = len(data)
                ard_list[1] = ard_list[2] + ard_list[3]
                ard_list[0] = ard_list[2] + ard_list[3] + ard_list[4]
                all_dict[name[0]] = ard_list
        # 生成fiter ou
        channel_banner_df = pd.read_excel('pop.xlsx')
        channel_df = channel_banner_df.iloc[:, 11].dropna(0)
        channel_list = list(channel_df)
        print(channel_list)
        banner_df = channel_banner_df.iloc[:, 22].dropna(0)
        banner_list = list(banner_df)
        banner_list.insert(0, "Total")
        print(banner_list)
        count = 0
        all_data = []
        for j, one_channel in enumerate(channel_list):
            next_count = ''
            for one_banner in banner_list:
                if j == 0:
                    count += 1
                    next_count = count
                else:  # 'Total' 'H/S' 'Hyper', 'Super', 'Mini'
                    if one_banner == "Total":
                        count += 1
                        next_count = count
                    else:
                        channel_type = all_dict[one_banner]  # 获取各个超市对应的数据
                        if one_channel == 'H/S':  # 如果名为“H/S”
                            if channel_type[1] == 0:
                                continue
                            if channel_type[1] != channel_type[0]:
                                count += 1
                                next_count = count
                            else:
                                for one_data in all_data:
                                    if one_data[0] == 'Total' and one_data[1] == one_banner:
                                        next_count = one_data[-1]
                                        break
                        elif one_channel == 'Hyper':
                            if channel_type[2] == 0:
                                continue
                            if channel_type[2] != channel_type[1]:
                                count += 1
                                next_count = count
                            else:
                                for one_data in all_data:
                                    if one_data[0] == 'H/S' and one_data[1] == one_banner:
                                        next_count = one_data[-1]
                                        break
                        elif one_channel == 'Super':

                            if channel_type[3] == 0:
                                continue
                            if channel_type[3] != channel_type[1]:
                                count += 1
                                next_count = count
                            else:
                                for one_data in all_data:
                                    if one_data[0] == 'H/S' and one_data[1] == one_banner:
                                        next_count = one_data[-1]
                                        break
                        elif one_channel == 'Mini':
                            if channel_type[4] == 0:
                                continue
                            if channel_type[4] != channel_type[0]:
                                count += 1
                                next_count = count
                            else:
                                for one_data in all_data:
                                    if one_data[0] == 'Mini' and one_data[1] == one_banner:
                                        next_count = one_data[-1]
                                        break
                        else:
                            count += 1
                            next_count = count
                lins_list = []
                lins_list.append(one_channel)
                lins_list.append(one_banner)
                tag = one_channel + one_banner
                lins_list.append(tag)
                lins_list.append(next_count)
                all_data.append(lins_list)
        df_ou = pd.DataFrame(data=all_data, columns=["Channel", "Banner", "Combination", "FilterNo"])
        df_ou.to_excel('fiter_byou.xlsx', index=False, sheet_name='fiterbyou')

    def summary_all_file(self):
        """把所有子表整理成一张表"""

        bybanner_df = pd.read_excel('测试banner_ou总和.xlsx', sheet_name='bybanner')  # 读取 测试路径总和.xlsx 生成的
        byou_df = pd.read_excel('测试banner_ou总和.xlsx', sheet_name='byou')  # 读取 测试路径总和.xlsx 生成的
        fitbanner_df = pd.read_excel('fiter_bybanner.xlsx',
                                     sheet_name='fiterbybanner')  # 读取 pop.xlsx 生成 测试去重excel.xlsx 生成的
        fitou_df = pd.read_excel('fiter_byou.xlsx', sheet_name='fiterbyou')  # 读取 pop.xlsx  生成的
        allpath_df = pd.read_excel('测试路径总和.xlsx', sheet_name='测试')  # 读取所有文件生产的
        pop_df = pd.read_excel('pop.xlsx')  # pop 文件

        # tablecopy_df = pd.read_excel('测试汇总设置表.xlsx',sheet_name='TableCopy')
        writer = pd.ExcelWriter("测试设置表.xlsx")
        bybanner_df.to_excel(writer, index=False, sheet_name='bybanner')
        byou_df.to_excel(writer, index=False, sheet_name='byou')
        fitbanner_df.to_excel(writer, index=False, sheet_name='fiterbybanner')
        fitou_df.to_excel(writer, index=False, sheet_name='fiterbyou')
        allpath_df.to_excel(writer, index=False, sheet_name='Dirs')
        pop_df.to_excel(writer, index=False, sheet_name='pop')
        # tablecopy_df.to_excel(writer,index=False,sheet_name="TableCopy")
        writer.save()
        writer.close()
        self.write_fiterline_table()  # 存筛选表

        # wb = openpyxl.load_workbook(name)

    def write_fiterline_table(self):
        """整合设计表"""

        colors = ['EEE5DE', 'FFE1FF', 'C1FFC1', 'C6E2FF', 'FAEBD7']  # 颜色编码
        font = Font('宋体', bold=True, color='FFFFFF')
        wb = openpyxl.load_workbook('测试汇总设置表.xlsx')
        ws = wb['TableCopy']
        # sheets1=wb.sheetnames()#获取sheet页
        rows = ws.rows
        all_data_list = []
        for i in rows:
            lins_list = []
            for one_i in i:
                lins_list.append(one_i.value)
            # print(lins_list)
            all_data_list.append(lins_list)
            # print(i)
        wb = openpyxl.load_workbook('测试设置表.xlsx')
        wb.create_sheet('TableCopy')  # 新建一个sheet
        ws = wb['TableCopy']
        for i in all_data_list:
            ws.append(i)
        fill = PatternFill("solid", fgColor="36648B")  # 标题
        zm = ['A', 'B', 'C', 'D', 'E', 'F']
        for one_zm in zm:
            ws['{}1'.format(one_zm)].fill = fill
            ws['{}1'.format(one_zm)].font = font
        color = colors[0]
        color_id = 0
        for i in range(2, len(all_data_list)):
            # print(colors[color_id])
            fill = PatternFill("solid", fgColor=colors[color_id])
            # print(fill)
            ws['A{}'.format(i)].fill = fill
            ws['B{}'.format(i)].fill = fill
            ws['C{}'.format(i)].fill = fill
            ws['D{}'.format(i)].fill = fill
            ws['E{}'.format(i)].fill = fill
            ws['F{}'.format(i)].fill = fill
            if ws['A{}'.format(i)].value != ws['A{}'.format(i + 1)].value:
                color_id += 1
                if color_id == 5:
                    color_id = 1
        wb.save('{}.xlsx'.format('测试设置表'))

    def run(self):
        self.get_bybanner_city()
        self.get_bybanner_fiter_no()
        self.get_byou_fiter_no()
        self.get_bybanner_ou_No()
        self.summary_all_file()




if __name__ == '__main__':
    a = GetFilter()
    a.run()
    # a = SaveInSql()
    # a.run()
    ###

    # a.get_bybanner_city()
    # a.get_bybanner_fiter_no()
    # a.get_byou_fiter_no()
    # a.get_bybanner_ou_No()
    # a.summary_all_file()
    # VisitMain()
    # a = Text()
    # a.get_fitter_list()
