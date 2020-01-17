from PIL import Image
import openpyxl
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import PatternFill

# img = Image.open('D:\杨伟强\整理\项目模块\动态可视化/1.png')
# print(img)


wb = openpyxl.Workbook()
ws = wb.create_sheet()

rows = [
    ('Batch 1', 'Batch 2'),
    (1.7, 1.8),
    (3.65, 3.55),
    (3.85, 3.5),

]

for row in rows:
    ws.append(row)

a = PatternFill(fill_type=None)
chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
# chart = BarChart()
# chart.add_data(values)
# ws.add_chart(chart, "E15")

# chart1.title = "Bar Chart"


data = Reference(ws, min_col=1, min_row=2, max_row=7, max_col=3)
# cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data)
# chart1.set_categories(cats)
# chart1.shape = 4
ws.add_chart(chart1, "A10")
wb.save('12.xlsx')
